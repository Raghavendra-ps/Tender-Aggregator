import os # Make sure 'os' is imported if not already
import re
import io
import shutil
import json
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple, Set
import datetime # Ensure datetime is imported
import logging
import subprocess
import sys
import time
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup

from fastapi import FastAPI, Request, Form, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse, FileResponse, PlainTextResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape
from starlette.datastructures import URL
from pydantic import BaseModel

# Playwright import is NOT needed here by dashboard for the new worker flow
# import base64 # NOT needed here by dashboard for the new worker flow
import httpx # For AI proxy

# --- Filter Engine Imports (Robust) ---
try:
    from filter_engine import run_filter, parse_tender_blocks_from_tagged_file, extract_tender_info_from_tagged_block # << INDIAN_STATES REMOVED
    filter_engine_available = True
except ImportError: 
    print("ERROR: Could not import 'filter_engine' components. Filter and AI data prep functionality may be affected.")
    # INDIAN_STATES = ["Error: State List Unavailable"] # This fallback is no longer needed here as we are not importing it.
    # If any part of dashboard.py *was* trying to use this imported INDIAN_STATES directly,
    # that part will now fail with a NameError, and you'd need to fix it
    # by using available_sites_main or loading from settings.json instead.
    # However, it's unlikely dashboard.py was using it for much, given the settings.json source.

    # Keep the dummy function definitions for filter_engine components if import fails:
    filter_engine_available = False
    def _log_filter_engine_error(message):
        logger_ref = globals().get('logger')
        if logger_ref and isinstance(logger_ref, logging.Logger):
            logger_ref.error(message)
        else:
            print(f"FILTER_ENGINE_ERROR: {message}")

    def run_filter(*args, **kwargs): 
        _log_filter_engine_error("CRITICAL ERROR: filter_engine.run_filter not imported."); 
        raise RuntimeError("filter_engine.run_filter not available.")
    def parse_tender_blocks_from_tagged_file(*args, **kwargs):
        _log_filter_engine_error("CRITICAL ERROR: filter_engine.parse_tender_blocks_from_tagged_file not imported.");
        raise RuntimeError("filter_engine.parse_tender_blocks_from_tagged_file not available.")
    def extract_tender_info_from_tagged_block(*args, **kwargs):
        _log_filter_engine_error("CRITICAL ERROR: filter_engine.extract_tender_info_from_tagged_block not imported.");
        raise RuntimeError("filter_engine.extract_tender_info_from_tagged_block not available.")

# --- MAIN_SITE_CONFIGS Import ---
available_sites_main: List[str] = ["Error: MAIN_SITE_CONFIGS not loaded"]
MAIN_SITE_CONFIGS: Dict[str, Any] = {}
try:
    from scrape import SITE_CONFIGS as MAIN_SITE_CONFIGS_IMPORT
    if isinstance(MAIN_SITE_CONFIGS_IMPORT, dict) and MAIN_SITE_CONFIGS_IMPORT:
        MAIN_SITE_CONFIGS = MAIN_SITE_CONFIGS_IMPORT
        available_sites_main = sorted(list(MAIN_SITE_CONFIGS.keys()))
        if not available_sites_main: print("ERROR: SITE_CONFIGS from scrape.py has no keys."); available_sites_main = ["Error: Main Config Empty"]
    else: print("ERROR: SITE_CONFIGS imported from scrape.py is empty or not a dict."); available_sites_main = ["Error: Main Config Invalid"]
except ImportError: print("ERROR: Could not import 'scrape' module for MAIN_SITE_CONFIGS.")
except Exception as e_main_configs: print(f"ERROR: Could not process SITE_CONFIGS from scrape.py: {e_main_configs}"); available_sites_main = ["Error: Main Config Import Failed"]

# --- URL_SUFFIX_ROT Import ---
URL_SUFFIX_ROT_DEFAULT = "?page=WebTenderStatusLists&service=page"
URL_SUFFIX_ROT = URL_SUFFIX_ROT_DEFAULT
try:
    from headless_rot_worker import URL_SUFFIX_ROT as URL_SUFFIX_FROM_WORKER
    URL_SUFFIX_ROT = URL_SUFFIX_FROM_WORKER
    print("INFO: Dashboard using URL_SUFFIX_ROT from 'headless_rot_worker.py'.")
except ImportError:
    print(f"WARNING: Could not import URL_SUFFIX_ROT from 'headless_rot_worker.py'. Using dashboard default: {URL_SUFFIX_ROT_DEFAULT}")
except Exception as e_url_suffix:
    print(f"WARNING: Error importing URL_SUFFIX_ROT from headless_rot_worker: {e_url_suffix}. Using dashboard default.")


# === NEW UNIFIED PATH CONFIGURATION ===
try:
    PROJECT_ROOT = Path(__file__).parent.resolve()
except NameError:
    PROJECT_ROOT = Path('.').resolve()

SITE_DATA_ROOT = PROJECT_ROOT / "site_data"

# ROT Specific Paths
ROT_DATA_DIR = SITE_DATA_ROOT / "ROT"
ROT_MERGED_SITE_SPECIFIC_DIR = ROT_DATA_DIR / "MergedSiteSpecific"
ROT_DETAIL_HTMLS_DIR = ROT_DATA_DIR / "DetailHtmls"
ROT_FINAL_GLOBAL_MERGED_DIR = ROT_DATA_DIR / "FinalGlobalMerged"
ROT_AI_ANALYSIS_DIR = ROT_DATA_DIR / "AI_Analysis"
ROT_FILTERED_RESULTS_DIR = ROT_DATA_DIR / "FilteredResults"

# REG Specific Paths
REG_DATA_DIR = SITE_DATA_ROOT / "REG"
REG_MERGED_SITE_SPECIFIC_DIR = REG_DATA_DIR / "MergedSiteSpecific"
REG_FINAL_GLOBAL_MERGED_DIR = REG_DATA_DIR / "FinalGlobalMerged"
REG_FILTERED_RESULTS_DIR = REG_DATA_DIR / "FilteredResults"

# TEMP Specific Paths
TEMP_DATA_DIR = SITE_DATA_ROOT / "TEMP"
TEMP_WORKER_RUNS_DIR = TEMP_DATA_DIR / "WorkerRuns"
TEMP_DEBUG_SCREENSHOTS_DIR_DASHBOARD = TEMP_DATA_DIR / "DebugScreenshots" / "dashboard"

# LOGS Path (Unified)
LOGS_BASE_DIR = PROJECT_ROOT / "LOGS"

# Other constants
TEMPLATES_DIR = PROJECT_ROOT / "templates" # Assuming templates is at project root
SETTINGS_FILE = PROJECT_ROOT / "settings.json"
SITE_CONTROLLER_SCRIPT_PATH = PROJECT_ROOT / "site_controller.py"
SCHEDULER_SETUP_SCRIPT_PATH = PROJECT_ROOT / "scheduler_setup.py"
FILTERED_TENDERS_FILENAME = "Filtered_Tenders.json"

AI_ANALYSIS_DATA_FILE = ROT_AI_ANALYSIS_DIR / "ai_analysis_data.json"

OPENWEBUI_API_BASE_URL_CONFIG = os.environ.get("OPENWEBUI_API_BASE_URL", "http://192.168.1.104:8080")
OPENWEBUI_API_KEY_CONFIG = os.environ.get("OPENWEBUI_API_KEY", "sk-your-key-here")
DEFAULT_OLLAMA_MODEL_ID_CONFIG = os.environ.get("DEFAULT_OLLAMA_MODEL_ID", "gemma3:12b")

# --- Create Directories ---
ROT_MERGED_SITE_SPECIFIC_DIR.mkdir(parents=True, exist_ok=True)
ROT_DETAIL_HTMLS_DIR.mkdir(parents=True, exist_ok=True)
ROT_FINAL_GLOBAL_MERGED_DIR.mkdir(parents=True, exist_ok=True)
ROT_AI_ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
ROT_FILTERED_RESULTS_DIR.mkdir(parents=True, exist_ok=True)

REG_MERGED_SITE_SPECIFIC_DIR.mkdir(parents=True, exist_ok=True)
REG_FINAL_GLOBAL_MERGED_DIR.mkdir(parents=True, exist_ok=True)
REG_FILTERED_RESULTS_DIR.mkdir(parents=True, exist_ok=True)

TEMP_WORKER_RUNS_DIR.mkdir(parents=True, exist_ok=True)
TEMP_DEBUG_SCREENSHOTS_DIR_DASHBOARD.mkdir(parents=True, exist_ok=True)

LOGS_BASE_DIR.mkdir(parents=True, exist_ok=True)
(LOGS_BASE_DIR / "regular_scraper").mkdir(parents=True, exist_ok=True)
(LOGS_BASE_DIR / "rot_worker").mkdir(parents=True, exist_ok=True)
# --- END Create Directories ---

# --- Logging Setup for Dashboard ---
log_formatter = logging.Formatter("%(asctime)s [%(levelname)s] (Dashboard) %(message)s", datefmt="%H:%M:%S")
log_handler_stream = logging.StreamHandler()
log_handler_stream.setFormatter(log_formatter)
dashboard_log_file = LOGS_BASE_DIR / "dashboard.log"
log_file_handler_dashboard = logging.FileHandler(dashboard_log_file, mode='a', encoding='utf-8')
log_file_handler_dashboard.setFormatter(log_formatter)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.hasHandlers():
    logger.addHandler(log_handler_stream)
    logger.addHandler(log_file_handler_dashboard)
    logger.propagate = False
# --- END Logging Setup for Dashboard ---

# --- DEFAULT_SETTINGS dictionary definition (ensure it's the one with post_submit) ---
DEFAULT_SETTINGS = {
  "global_scraper_settings": {
    "scrape_py_limits": {"max_list_pages": 100, "retries": 3}, "concurrency": {"list_pages": 5, "detail_pages": 2}, "timeouts": {"page_load": 75000, "detail_page": 75000},
    "rot_scrape_limits": {"max_list_pages": 50, "retries": 2},
    "rot_concurrency": {"detail_processing": 2},
    "rot_timeouts": {
        "page_load": 75000, "detail_page": 60000, "download": 180000,
        "pagination_load_wait": 7000, "element_wait": 20000, "post_submit": 30000 # Has post_submit
     }
  }, "retention": {"enabled": False,"days": 30 },
  "scheduler": { "main_enabled": False, "main_frequency": "daily", "main_time": "02:00", "main_day_weekly": "7", "main_day_monthly": "1",
                 "rot_default_status_value": "5" },
  "site_configurations": {}
}

# These might not be needed globally anymore if worker_settings are constructed specifically
# PAGE_LOAD_TIMEOUT_DEFAULT = DEFAULT_SETTINGS["global_scraper_settings"]["timeouts"]["page_load"]
# ELEMENT_WAIT_TIMEOUT_DEFAULT = DEFAULT_SETTINGS["global_scraper_settings"]["rot_timeouts"].get("element_wait", 20000)

# --- App and Jinja Setup ---
app = FastAPI(title="TenFin Tender Dashboard")
templates: Optional[Jinja2Templates] = None
if TEMPLATES_DIR.is_dir(): # TEMPLATES_DIR is now PROJECT_ROOT / "templates"
    try:
        jinja_env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)), autoescape=select_autoescape(['html', 'xml']))
        jinja_env.filters['neat_format'] = format_text_neatly # Assuming format_text_neatly is defined
        templates = Jinja2Templates(env=jinja_env)
        logger.info("Jinja2 configured with new templates path.")
    except Exception as e:
        logger.error(f"Jinja2 init failed with new templates path: {e}", exc_info=True)
        templates = None
else:
    templates = None
    logger.error(f"Templates dir '{TEMPLATES_DIR}' not found. UI will be broken.")
def load_settings() -> Dict:
    if not SETTINGS_FILE.is_file():
        logger.warning(f"Settings file not found at {SETTINGS_FILE}. Creating with defaults.")
        try:
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f: json.dump(DEFAULT_SETTINGS, f, indent=2)
            logger.info(f"Created default settings file: {SETTINGS_FILE}")
            return DEFAULT_SETTINGS.copy()
        except Exception as e: logger.error(f"Could not create default settings: {e}"); return DEFAULT_SETTINGS.copy()
    try:
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f: settings = json.load(f)
        needs_resave = False
        for main_key, default_value in DEFAULT_SETTINGS.items():
            if main_key not in settings:
                settings[main_key] = default_value.copy(); needs_resave = True
            elif isinstance(default_value, dict):
                if not isinstance(settings.get(main_key), dict):
                    settings[main_key] = default_value.copy(); needs_resave = True
                else:
                    for sub_key, sub_default_value in default_value.items():
                        if sub_key not in settings[main_key]:
                            settings[main_key][sub_key] = sub_default_value; needs_resave = True
                        elif isinstance(sub_default_value, dict):
                             if not isinstance(settings[main_key].get(sub_key), dict):
                                 settings[main_key][sub_key] = sub_default_value.copy(); needs_resave = True
                             else:
                                 for deep_key, deep_default in sub_default_value.items():
                                     if deep_key not in settings[main_key][sub_key]:
                                         settings[main_key][sub_key][deep_key] = deep_default; needs_resave = True
        if "scheduler" in settings and isinstance(settings["scheduler"], dict):
            if "paused" in settings["scheduler"]: del settings["scheduler"]["paused"]; needs_resave = True
            for old_rot_key in ["rot_enabled", "rot_frequency", "rot_time", "rot_day_weekly", "rot_day_monthly"]:
                if old_rot_key in settings["scheduler"]: del settings["scheduler"][old_rot_key]; needs_resave = True
            if "rot_default_status_value" not in settings["scheduler"]:
                 settings["scheduler"]["rot_default_status_value"] = DEFAULT_SETTINGS["scheduler"]["rot_default_status_value"]
                 needs_resave = True
        return settings
    except json.JSONDecodeError as e: logger.error(f"Error decoding {SETTINGS_FILE}: {e}. Defaults returned."); return DEFAULT_SETTINGS.copy()
    except Exception as e: logger.error(f"Error reading {SETTINGS_FILE}: {e}. Defaults returned.", exc_info=True); return DEFAULT_SETTINGS.copy()

# In dashboard.py
# Ensure BeautifulSoup, re, Path, logger, urlparse, urljoin are available

def parse_rot_summary_html(html_file_path: Path, site_key_original: str) -> Dict[str, Any]: # Added site_key_original
    log_prefix_parser = f"[ROTParse-{html_file_path.stem}]"
    logger.info(f"{log_prefix_parser} Starting parsing of: {html_file_path.name}")
    extracted_data = {
        "original_filename": html_file_path.name,
        "page_title": site_key_original, # MODIFIED: Default to original site_key
        # "key_details": {}, # REMOVED: No longer extracting general info into key_details for display
        "sections": [],
        "error_message": None
    }

    if not html_file_path.is_file():
        extracted_data["error_message"] = "Summary HTML file not found on server."
        logger.error(f"{log_prefix_parser} {extracted_data['error_message']}")
        return extracted_data

    try:
        with open(html_file_path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        soup = BeautifulSoup(content, 'html.parser')

        # Page Title: Use the passed site_key_original as the primary title.
        # The full title from the HTML (like "eProcurement System...") can be a sub-heading if needed.
        # For now, we set page_title directly to site_key_original.

        main_content_wrapper = soup.find('form', id='bidSummaryForm')
        if not main_content_wrapper: main_content_wrapper = soup.find('div', class_='border')
        if not main_content_wrapper: main_content_wrapper = soup # Fallback

        # REMOVED: Logic for extracting the first info table into "key_details"
        # That information is already on the view_rot.html list page.

        # --- Extract Sections based on <td class="section_head"> or other cues ---
        section_headers_to_find = [
            "Bids List",
            "Technical Bid Opening Summary",
            "Technical Evaluation Summary Details",
            "Finance Bid Opening Summary",
            "Financial Evaluation Bid List",
            "Finance Evaluation Summary Details"
        ]
        
        if main_content_wrapper:
            all_tables_in_content = main_content_wrapper.find_all('table', recursive=True)
            processed_tables_for_sections = set()

            # First, try to find tables by their explicit section head class
            section_head_tds = main_content_wrapper.find_all('td', class_='section_head')
            
            for header_td in section_head_tds:
                section_title = header_td.get_text(strip=True)
                
                # Find the table this header_td belongs to or the one immediately following its row.
                current_search_element = header_td.find_parent('tr')
                if not current_search_element: current_search_element = header_td # If header is not in a TR
                
                associated_table = None
                # Try to find the table as a sibling of the row containing the header, or parent table
                limit_scan = 0
                temp_element = current_search_element
                while temp_element and limit_scan < 5:
                    candidate_table = temp_element.find_next_sibling('table', class_=re.compile(r'(table_list|list_table|tablebg)', re.I))
                    if candidate_table:
                        associated_table = candidate_table
                        break
                    # Check if current_search_element's parent IS the table we want (header is first row)
                    if temp_element.parent and temp_element.parent.name == 'table' and \
                       'table_list' in temp_element.parent.get('class',[]):
                        associated_table = temp_element.parent
                        break
                    temp_element = temp_element.parent
                    limit_scan +=1
                
                if not associated_table and header_td: # Last resort if structure is flatter
                    associated_table = header_td.find_next('table', class_=re.compile(r'(table_list|list_table|tablebg)', re.I))


                if associated_table and associated_table not in processed_tables_for_sections:
                    # MODIFICATION: Remove duplicate header row if present
                    first_tr = associated_table.find('tr')
                    if first_tr:
                        first_td_or_th = first_tr.find(['td', 'th'])
                        if first_td_or_th and section_title.lower() in first_td_or_th.get_text(strip=True).lower() and \
                           ('section_head' in first_td_or_th.get('class', []) or len(first_tr.find_all(['td','th'])) == 1) :
                            logger.debug(f"{log_prefix_parser} Decomposing duplicate header row for section: '{section_title}'")
                            first_tr.decompose() # Remove the row containing the duplicate header

                    # MODIFICATION: Remove "Document : None" rows
                    rows_to_remove = []
                    for row in associated_table.find_all('tr'):
                        cells = row.find_all('td')
                        if len(cells) >= 2: # Assuming Key-Value structure
                            key_text = cells[0].get_text(strip=True)
                            value_text = cells[1].get_text(strip=True)
                            if "document" in key_text.lower() and value_text.lower() == "none":
                                rows_to_remove.append(row)
                                logger.debug(f"{log_prefix_parser} Marking 'Document : None' row for removal in section '{section_title}'.")
                    for row_to_remove in rows_to_remove:
                        row_to_remove.decompose()
                    
                    # Basic cleanup
                    for s_tag in associated_table.find_all(['script', 'style', 'link']): s_tag.decompose()
                    for a_tag in associated_table.find_all('a'): a_tag.replace_with(a_tag.get_text(strip=True))


                    extracted_data["sections"].append({
                        "title": section_title,
                        "html_content": str(associated_table)
                    })
                    processed_tables_for_sections.add(associated_table)
                    logger.debug(f"{log_prefix_parser} Extracted section by table header: '{section_title}'")
                elif not associated_table:
                     logger.warning(f"{log_prefix_parser} Could not find associated table for section header: '{section_title}'")


        # Fallback if no specific sections were found
        if not extracted_data["sections"]:
            logger.warning(f"{log_prefix_parser} No specific sections extracted. Falling back to full content.")
            body_content = soup.body if soup.body else soup
            if body_content:
                 for s_tag in body_content.find_all(['script', 'style', 'link']): s_tag.decompose()
                 for a_tag in body_content.find_all('a'): a_tag.replace_with(a_tag.get_text(strip=True))
                 extracted_data["sections"].append({
                        "title": "Full Cleaned Summary Content", # Changed title
                        "html_content": str(body_content)
                    })
            else:
                extracted_data["error_message"] = "Could not extract any content sections from the summary HTML."
                logger.warning(f"{log_prefix_parser} No sections or body content found to extract.")

    except Exception as e:
        logger.error(f"{log_prefix_parser} Error during BeautifulSoup parsing of {html_file_path.name}: {e}", exc_info=True)
        extracted_data["error_message"] = f"An error occurred while parsing the summary content: {str(e)}"
    
    logger.info(f"{log_prefix_parser} Parsing complete. Page Title: '{extracted_data['page_title']}', Sections: {len(extracted_data['sections'])}")
    return extracted_data

def save_settings(settings_data: Dict) -> bool:
    try:
        if "scheduler" in settings_data and isinstance(settings_data["scheduler"], dict):
            if "paused" in settings_data["scheduler"]:
                del settings_data["scheduler"]["paused"]; logger.info("Removed 'paused' key from scheduler settings before saving.")
            for key in ["rot_enabled", "rot_frequency", "rot_time", "rot_day_weekly", "rot_day_monthly"]:
                if key in settings_data["scheduler"]:
                    del settings_data["scheduler"][key]
            if "rot_default_status_value" not in settings_data["scheduler"]:
                 settings_data["scheduler"]["rot_default_status_value"] = DEFAULT_SETTINGS["scheduler"]["rot_default_status_value"]
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f: json.dump(settings_data, f, indent=2, ensure_ascii=False)
        logger.info(f"Settings saved successfully to {SETTINGS_FILE}"); return True
    except Exception as e: logger.error(f"Error saving settings to {SETTINGS_FILE}: {e}"); return False

def format_text_neatly(text: Optional[str], min_len_for_titlecase=8) -> str:
    if not text or not isinstance(text, str): return ""
    cleaned_text = re.sub(r'[^\w\s.,()/\-]', ' ', text); cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
    if not cleaned_text: return ""
    if len(cleaned_text) >= min_len_for_titlecase and cleaned_text.isupper() and ' ' not in cleaned_text and cleaned_text.isalnum(): return cleaned_text
    if ' ' in cleaned_text and cleaned_text.isupper(): return cleaned_text.title()
    formatted_text = cleaned_text[0].upper() + cleaned_text[1:].lower(); formatted_text = re.sub(r'(?<=[.!?]\s)([a-z])', lambda m: m.group(1).upper(), formatted_text)
    return formatted_text

app = FastAPI(title="TenFin Tender Dashboard")
templates: Optional[Jinja2Templates] = None
if TEMPLATES_DIR.is_dir():
    try:
        jinja_env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)), autoescape=select_autoescape(['html', 'xml']))
        jinja_env.filters['neat_format'] = format_text_neatly; templates = Jinja2Templates(env=jinja_env); logger.info("Jinja2 configured.")
    except Exception as e: logger.error(f"Jinja2 init failed: {e}", exc_info=True); templates = None
else: templates = None; logger.error(f"Templates dir '{TEMPLATES_DIR}' not found. UI broken.")

def _validate_subdir(subdir: str, base_dir: Path) -> Path: 
    if not subdir or not isinstance(subdir, str): raise HTTPException(400, "Invalid subdir name.")
    if ".." in subdir or subdir.startswith(("/", "\\")) or any(c in subdir for c in ['<', '>', ':', '"', '|', '?', '*']): raise HTTPException(400, "Invalid subdir format.")
    if not base_dir.is_dir(): raise HTTPException(500, "Server base dir error.")
    try:
        subdir_path_part = Path(subdir)
        if subdir_path_part.is_absolute() or subdir_path_part.drive: raise HTTPException(400, "Subdir must be relative.")
        full_path = (base_dir / subdir_path_part).resolve(); resolved_base_dir = base_dir.resolve()
        if not str(full_path).startswith(str(resolved_base_dir)): raise HTTPException(400, "Path traversal detected.")
        if resolved_base_dir != full_path.parent and resolved_base_dir not in full_path.parents: raise HTTPException(400, "Invalid path nesting.")
    except ValueError as ve: raise HTTPException(400, f"Invalid path: {ve}")
    except HTTPException: raise
    except Exception as e: logger.error(f"Path validation error for '{subdir}': {e}", exc_info=True); raise HTTPException(500, "Internal path error.")
    return full_path

def write_dashboard_status(run_dir: Path, status_message: str): # Renamed to avoid conflict if importing from worker
    # Uses the dashboard's logger
    log = logger 
    try:
        status_file = run_dir / "status.txt"
        status_file.write_text(status_message, encoding='utf-8')
        log.info(f"Dashboard: Status updated to '{status_message}' in {status_file}")
    except Exception as e: 
        log.error(f"Dashboard: Error writing status '{status_message}' to {run_dir}: {e}")

def _clean_path_component(component: str) -> str: # <<<< DEFINITION HERE
    """
    Removes potentially unsafe characters from a path component.
    Allows alphanumeric, underscore, hyphen, period.
    Prevents components from being just "." or "..".
    """
    if not isinstance(component, str):
        if 'logger' in globals() and isinstance(globals()['logger'], logging.Logger):
            logger.warning(f"Attempted to clean non-string path component: {type(component)}")
        return ""
    cleaned = re.sub(r'[^\w\-._]', '', component)
    if cleaned == "." or cleaned == "..":
        if 'logger' in globals() and isinstance(globals()['logger'], logging.Logger):
            logger.warning(f"Path component cleaned to unsafe value '{cleaned}', returning empty string.")
        return ""
    return cleaned

def cleanup_old_filtered_results():
    logger.info("Running cleanup for old filtered results (REGULAR and ROT)...")
    settings = load_settings()
    retention_settings = settings.get("retention", {}) # Get the retention sub-dictionary

    if not retention_settings.get("enabled", False):
        logger.info("Retention cleanup disabled in settings.")
        return

    try:
        # Ensure 'days' is valid, with a fallback if key is missing or invalid
        days_to_keep_str = str(retention_settings.get("days", "30")) # Get as string first
        if not days_to_keep_str.isdigit():
            logger.warning(f"Invalid 'days' value '{days_to_keep_str}' in retention settings. Using default 30.")
            days_to_keep = 30
        else:
            days_to_keep = int(days_to_keep_str)
        
        days_to_keep = max(1, min(days_to_keep, 3650)) # Keep within a reasonable range (1 day to 10 years)
        
        cutoff_timestamp = time.time() - (days_to_keep * 24 * 60 * 60)
        
        # MODIFIED: Use new path constants for filtered results
        paths_to_clean_map = {
            "REGULAR": REG_FILTERED_RESULTS_DIR,
            "ROT": ROT_FILTERED_RESULTS_DIR
        }

        for data_type_label, base_to_clean in paths_to_clean_map.items():
            deleted_count, kept_count, error_count = 0, 0, 0
            logger.info(f"Cleaning {data_type_label} filtered tenders older than {days_to_keep} days in: {base_to_clean}")

            if not base_to_clean.is_dir():
                logger.warning(f"Directory for {data_type_label} filtered results not found: '{base_to_clean}'. Skipping cleanup for this type.")
                continue

            for item in base_to_clean.iterdir():
                if item.is_dir(): # We expect subdirectories named after filters
                    try:
                        item_mtime = item.stat().st_mtime
                        if item_mtime < cutoff_timestamp:
                            logger.info(f"Deleting old {data_type_label} filtered set: {item.name} (Modified: {datetime.datetime.fromtimestamp(item_mtime).strftime('%Y-%m-%d %H:%M:%S')})")
                            shutil.rmtree(item)
                            deleted_count += 1
                        else:
                            kept_count += 1
                    except Exception as e:
                        logger.error(f"Error processing directory '{item.name}' (Type: {data_type_label}) during cleanup: {e}", exc_info=False) # exc_info=False to avoid too much noise for simple file errors
                        error_count += 1
            
            logger.info(f"Retention cleanup for {data_type_label} finished. Deleted: {deleted_count}, Kept: {kept_count}, Errors: {error_count}")

    except ValueError as ve: # This might catch int conversion if not handled above, but a more specific error
        logger.error(f"Invalid 'days' value in retention settings, causing ValueError: {ve}")
    except Exception as e:
        logger.error(f"Unexpected error during retention cleanup: {e}", exc_info=True)

def get_rot_site_config(site_key_to_find: str) -> Optional[Dict[str, str]]:
    settings = load_settings()
    all_site_configs = settings.get("site_configurations", {})
    if site_key_to_find in all_site_configs:
        main_config = all_site_configs[site_key_to_find]
        domain_from_settings = main_config.get("domain")

        if domain_from_settings:
            cleaned_domain_for_search_url = domain_from_settings.rstrip('/?')
         
            rot_search_url = f"{cleaned_domain_for_search_url}{URL_SUFFIX_ROT}"

            parsed_cleaned_domain = urlparse(cleaned_domain_for_search_url)
            REG_FINAL_GLOBAL_MERGED_DIR_for_links = parsed_cleaned_domain.path
            if not REG_FINAL_GLOBAL_MERGED_DIR_for_links: REG_FINAL_GLOBAL_MERGED_DIR_for_links = "/"
            elif not REG_FINAL_GLOBAL_MERGED_DIR_for_links.endswith('/'): REG_FINAL_GLOBAL_MERGED_DIR_for_links += '/'
            base_url_for_resolving_links = f"{parsed_cleaned_domain.scheme}://{parsed_cleaned_domain.netloc}{REG_FINAL_GLOBAL_MERGED_DIR_for_links}"

            if ('nicgep' in base_url_for_resolving_links.lower() or 'eprocure.gov.in' in base_url_for_resolving_links.lower()):
                if not base_url_for_resolving_links.endswith('app/'):
                    path_segments = [s for s in parsed_cleaned_domain.path.split('/') if s]
                    if not path_segments or path_segments[-1].lower() != 'app':
                         base_url_for_resolving_links = urljoin(base_url_for_resolving_links, 'app/')

            logger.debug(f"Derived ROT config for '{site_key_to_find}': Search URL='{rot_search_url}', Base URL for links='{base_url_for_resolving_links}'")
            return {"main_search_url": rot_search_url, "base_url": base_url_for_resolving_links}
        else:
            logger.error(f"Domain not found for site_key: {site_key_to_find} in settings.")
    else:
        logger.error(f"Site key '{site_key_to_find}' not found in site_configurations.")
    return None

async def _globally_merge_rot_site_files(
    input_site_specific_merged_dir: Path, # e.g., ROT_MERGED_SITE_SPECIFIC_DIR
    output_final_global_dir: Path,        # e.g., ROT_FINAL_GLOBAL_MERGED_DIR
    data_type: str = "rot"                # Primarily for "rot" in this context
) -> Tuple[int, Optional[Path]]:
    """
    Merges site-specific tagged data files into a single global file.
    Called by dashboard endpoints.
    """
    log_prefix = f"DASHBOARD GLOBAL MERGE ({data_type.upper()})" # Differentiate from site_controller's merge
    logger.info(f"--- {log_prefix}: Starting Merge of All Site-Specific Files ---")
    logger.info(f"{log_prefix}: Input directory: {input_site_specific_merged_dir}")
    logger.info(f"{log_prefix}: Output directory: {output_final_global_dir}")

    if not input_site_specific_merged_dir.is_dir():
        logger.warning(f"{log_prefix}: Input directory for site-specific merged files not found: {input_site_specific_merged_dir}. No global merge will occur.")
        return 0, None

    file_pattern: str
    output_filename_prefix: str

    if data_type == "rot":
        file_pattern = "Merged_ROT_*.txt"
        output_filename_prefix = "Final_ROT_Tender_List_"
    elif data_type == "regular": # If you ever use this for regular data via dashboard
        file_pattern = "Merged_*.txt" # Assumes regular files are Merged_SITEKEY_DATE.txt
        output_filename_prefix = "Final_Tender_List_"
    else:
        logger.error(f"{log_prefix}: Invalid data_type '{data_type}' specified for global merge.")
        return 0, None

    site_merged_files = list(input_site_specific_merged_dir.glob(file_pattern))
    if not site_merged_files:
        logger.info(f"{log_prefix}: No '{file_pattern}' files found in {input_site_specific_merged_dir} for global merge.")
        return 0, None

    logger.info(f"{log_prefix}: Found {len(site_merged_files)} site-specific {data_type.upper()} files for consolidation.")
    global_timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")

    # Ensure the output directory exists
    output_final_global_dir.mkdir(parents=True, exist_ok=True)
    global_final_output_path = output_final_global_dir / f"{output_filename_prefix}{global_timestamp}.txt"

    total_globally_unique_tender_blocks = 0
    seen_global_tender_hashes: Set[int] = set() # To store hashes of unique blocks

    try:
        with open(global_final_output_path, "w", encoding="utf-8") as global_outfile:
            for site_file_path in site_merged_files:
                logger.info(f"{log_prefix}: Processing site-specific file: {site_file_path.name}")
                try:
                    site_file_content = site_file_path.read_text(encoding="utf-8", errors="replace").strip()
                    if not site_file_content:
                        logger.debug(f"{log_prefix}: File {site_file_path.name} is empty. Skipping.")
                        continue

                    # Split into blocks based on "--- TENDER END ---"
                    # Then reconstruct each block to ensure consistent formatting and include the end marker.
                    raw_blocks_with_delimiter = site_file_content.split("--- TENDER END ---")
                    tender_blocks_from_site_file: List[str] = []

                    for block_segment in raw_blocks_with_delimiter:
                        # A block must contain "--- TENDER START ---"
                        if "--- TENDER START ---" in block_segment:
                            # Find the actual start of the "--- TENDER START ---" marker
                            start_index = block_segment.find("--- TENDER START ---")
                            if start_index != -1:
                                # Reconstruct block: from "--- TENDER START ---" to end of segment, then add delimiter
                                reconstructed_block = block_segment[start_index:].strip() + "\n--- TENDER END ---"
                                tender_blocks_from_site_file.append(reconstructed_block.strip()) # Strip again to be sure

                    for full_block_text in tender_blocks_from_site_file:
                        if not full_block_text: # Should not happen if reconstruction is correct
                            continue
                        block_hash_global = hash(full_block_text)
                        if block_hash_global not in seen_global_tender_hashes:
                            global_outfile.write(full_block_text + "\n\n") # Two newlines between blocks
                            seen_global_tender_hashes.add(block_hash_global)
                            total_globally_unique_tender_blocks += 1
                except Exception as e_proc_site_file:
                    logger.error(f"{log_prefix}: Error processing content of site-specific file {site_file_path.name}: {e_proc_site_file}", exc_info=True) # Added exc_info

        if total_globally_unique_tender_blocks > 0:
            logger.info(f"{log_prefix}: ✅ Merged {total_globally_unique_tender_blocks} unique {data_type.upper()} tender blocks to: {global_final_output_path}")
        else:
            logger.info(f"{log_prefix}: No unique tender blocks found to merge into {global_final_output_path}. Output file may be empty or not created if no source blocks.")
            # Optionally delete the empty global file if no blocks were written
            if total_globally_unique_tender_blocks == 0 and global_final_output_path.exists():
                try:
                    global_final_output_path.unlink()
                    logger.info(f"{log_prefix}: Removed empty global output file: {global_final_output_path}")
                except OSError as e_del_empty:
                    logger.warning(f"{log_prefix}: Could not remove empty global output file {global_final_output_path}: {e_del_empty}")
                return 0, None # Indicate no file was effectively created

        return total_globally_unique_tender_blocks, global_final_output_path
    except Exception as e_global_merge:
        logger.error(f"{log_prefix}: Fatal error during global merge of {data_type.upper()} tenders: {e_global_merge}", exc_info=True)
        # Return current count (might be >0 if error happened mid-write) and None for path
        return total_globally_unique_tender_blocks, None

# === ROUTE HANDLERS ===

@app.get("/", response_class=HTMLResponse, name="homepage")
async def homepage(request: Request):
    if not templates: return HTMLResponse("Template engine error", status_code=503)
    subdirs_regular: List[str] = []
    # Use new constant for regular filtered results directory
    if REG_FILTERED_RESULTS_DIR.is_dir(): # MODIFIED
        try:
            subdirs_regular = sorted([
                d.name for d in REG_FILTERED_RESULTS_DIR.iterdir() # MODIFIED
                if d.is_dir() and not d.name.startswith('.')
            ])
        except OSError as e:
            logger.error(f"Error listing regular filtered results from {REG_FILTERED_RESULTS_DIR}: {e}") # MODIFIED
    else:
        logger.warning(f"Regular filtered results directory not found: '{REG_FILTERED_RESULTS_DIR}'") # MODIFIED

    subdirs_rot: List[str] = []
    # Use new constant for ROT filtered results directory
    if ROT_FILTERED_RESULTS_DIR.is_dir(): # MODIFIED
        try:
            subdirs_rot = sorted([
                d.name for d in ROT_FILTERED_RESULTS_DIR.iterdir() # MODIFIED
                if d.is_dir() and not d.name.startswith('.')
            ])
        except OSError as e:
            logger.error(f"Error listing ROT filtered results from {ROT_FILTERED_RESULTS_DIR}: {e}") # MODIFIED
    else:
        logger.warning(f"ROT filtered results directory not found: '{ROT_FILTERED_RESULTS_DIR}'") # MODIFIED
        
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "subdirs_eproc": subdirs_regular, # This key is used by index.html for regular
            "subdirs_rot": subdirs_rot      # This key is used by index.html for ROT
        }
    )

@app.get("/settings", name="settings_page", response_class=HTMLResponse)
async def settings_page(request: Request): 
    if not templates: return HTMLResponse("Template engine error", status_code=503)
    current_settings = load_settings()
    return templates.TemplateResponse("settings.html", {"request": request, "settings": current_settings })

@app.get("/rot-manual-scrape", name="rot_manual_scrape_page", response_class=HTMLResponse)
async def rot_manual_scrape_page(request: Request):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine not configured.")
    settings = load_settings()
    site_configs = settings.get('site_configurations', {})
    sites_available_for_rot = site_configs
    return templates.TemplateResponse("rot_manual_scrape.html", {
        "request": request,
        "sites_for_rot": sites_available_for_rot,
        "settings": settings
    })

@app.get("/rot-ai-analysis", name="rot_ai_analysis_page", response_class=HTMLResponse)
async def rot_ai_analysis_page_route(request: Request):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine not configured.")
    current_settings = load_settings() 
    return templates.TemplateResponse("rot_ai_analysis_dashboard.html", {
        "request": request,
        "settings": current_settings 
    })

@app.get("/view/{data_type}/{subdir}", name="view_tenders_typed", response_class=HTMLResponse)
async def view_tenders_typed(request: Request, data_type: str, subdir: str):
    if not templates: raise HTTPException(status_code=503, detail="Template engine error.") # Added status_code
    tenders: List[Dict[str, Any]] = []
    
    base_data_dir_for_view: Path # Define type for clarity

    if data_type == "rot":
        base_data_dir_for_view = ROT_FILTERED_RESULTS_DIR # MODIFIED
    elif data_type == "regular":
        base_data_dir_for_view = REG_FILTERED_RESULTS_DIR # MODIFIED
    else:
        raise HTTPException(status_code=400, detail="Invalid data type specified.") # Added status_code

    try:
        # _validate_subdir ensures subdir is safe and resolves it against base_data_dir_for_view
        subdir_path = _validate_subdir(subdir, base_dir=base_data_dir_for_view)
        file_path = subdir_path / FILTERED_TENDERS_FILENAME # FILTERED_TENDERS_FILENAME is "Filtered_Tenders.json"
    except HTTPException as e_val: # Catch validation specific errors
        logger.warning(f"Validation error for view '{subdir}' type '{data_type}': {e_val.detail}")
        raise e_val # Re-raise the validation HTTPException
    except Exception as e_path: # Catch other path construction errors
        logger.error(f"Unexpected path error view '{subdir}' type '{data_type}': {e_path}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal path error.") # Added status_code

    if not subdir_path.is_dir():
        logger.warning(f"Filter set directory not found: {subdir_path}")
        raise HTTPException(status_code=404, detail=f"Filter set '{subdir}' (Type: {data_type}) not found.") # Added status_code
    
    if not file_path.is_file():
        logger.warning(f"Data file '{FILTERED_TENDERS_FILENAME}' missing in '{subdir_path}'. Displaying empty list.")
        # tenders will remain empty, which is fine; template should handle empty list
    else:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                tenders_data = json.load(f)
            if not isinstance(tenders_data, list):
                logger.error(f"Invalid data format (not a list) in {file_path}. Displaying empty list.")
                # tenders will remain empty
            else:
                tenders = tenders_data
        except json.JSONDecodeError as e_json_dec:
            logger.error(f"Error parsing JSON from {file_path}: {e_json_dec}. Displaying empty list.")
            # tenders will remain empty
        except Exception as e_load: # Catch other file reading errors
            logger.error(f"Error loading data for '{subdir}' (Type: {data_type}) from {file_path}: {e_load}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error loading data for '{subdir}'.") # Added status_code
            
    subdir_display_name = subdir.replace('_Results', '').replace('_Tenders','').replace('_', ' ')
    template_name = "view_rot.html" if data_type == "rot" else "view.html"
    
    return templates.TemplateResponse(
        template_name,
        {
            "request": request,
            "subdir": subdir,
            "subdir_display_name": subdir_display_name,
            "tenders": tenders,
            "data_type": data_type
        }
    )

@app.get("/tender/{data_type}/{subdir}/{tender_id}", name="view_tender_detail_typed", response_class=HTMLResponse)
async def view_tender_detail_typed(request: Request, data_type: str, subdir: str, tender_id: str):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine error.") # Added status_code
    
    found_tender: Optional[Dict[str, Any]] = None
    base_data_dir_for_detail: Path # Define type

    if data_type == "rot":
        base_data_dir_for_detail = ROT_FILTERED_RESULTS_DIR # MODIFIED
    elif data_type == "regular":
        base_data_dir_for_detail = REG_FILTERED_RESULTS_DIR # MODIFIED
    else:
        raise HTTPException(status_code=400, detail="Invalid data type specified.") # Added status_code

    try:
        subdir_path = _validate_subdir(subdir, base_dir=base_data_dir_for_detail)
        file_path = subdir_path / FILTERED_TENDERS_FILENAME
    except HTTPException as e_val: # Catch validation specific errors
        logger.warning(f"Validation error for detail '{subdir}/{tender_id}' type '{data_type}': {e_val.detail}")
        raise e_val # Re-raise the validation HTTPException
    except Exception as e_path: # Catch other path construction errors
        logger.error(f"Unexpected path error for detail '{subdir}/{tender_id}' type '{data_type}': {e_path}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal path error for tender detail.")

    if not subdir_path.is_dir():
        logger.warning(f"Filter set directory not found for detail: {subdir_path}")
        raise HTTPException(status_code=404, detail=f"Filter set '{subdir}' (Type: {data_type}) not found.")
    
    if not file_path.is_file():
        logger.warning(f"Data file '{FILTERED_TENDERS_FILENAME}' missing for detail in '{subdir_path}'")
        raise HTTPException(status_code=404, detail=f"Data file missing for '{subdir}' (Type: {data_type}).")

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            tenders_list = json.load(f)
        
        if not isinstance(tenders_list, list):
            logger.error(f"Invalid data format (not a list) in {file_path} for tender detail.")
            raise ValueError("Invalid data format in tender list file.")

        for tender_item in tenders_list:
            if isinstance(tender_item, dict):
                # MODIFIED: Use specific ID keys based on data_type
                id_to_check = None
                if data_type == "rot":
                    id_to_check = tender_item.get("rot_tender_id")
                elif data_type == "regular":
                    # Prefer primary_tender_id, fall back to detail_tender_id if needed
                    id_to_check = tender_item.get("primary_tender_id")
                    if not id_to_check or id_to_check == "N/A": # Check for "N/A" as well
                        id_to_check = tender_item.get("detail_tender_id")
                
                if id_to_check == tender_id and id_to_check != "N/A": # Ensure we don't match on "N/A"
                    found_tender = tender_item
                    break
        
        if not found_tender:
            logger.warning(f"Tender ID '{tender_id}' not found in '{subdir_path}' (Type: {data_type}).")
            raise HTTPException(status_code=404, detail=f"Tender ID '{tender_id}' not found in '{subdir}' (Type: {data_type}).")

    except (json.JSONDecodeError, ValueError) as e_parse: # More specific error catching
        logger.error(f"Error reading/parsing {file_path} for tender detail: {e_parse}")
        raise HTTPException(status_code=500, detail=f"Error reading data for '{subdir}'.")
    except HTTPException: # Re-raise HTTPExceptions from above (like 404 if tender not found)
        raise
    except Exception as e_load: # Catch other unexpected errors
        logger.error(f"Unexpected error loading detail for '{subdir}/{tender_id}' (Type: {data_type}) from {file_path}: {e_load}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error loading tender details.")

    template_name = "tender_detail_rot.html" if data_type == "rot" else "tender_detail.html"
    
    # Prepare context for the template
    context = {
        "request": request,
        "tender": found_tender, # This will be the main tender dictionary
        "subdir": subdir,
        "data_type": data_type
    }
    # Specifically for rot_tender_detail.html, pass the tender again under 'rot_tender' if you prefer
    # Or just use 'tender' in both templates consistently.
    if data_type == "rot":
        context["rot_tender"] = found_tender # For explicit use in rot_tender_detail.html if desired

    return templates.TemplateResponse(template_name, context)




@app.get("/download/{data_type}/{subdir}", name="download_tender_excel_typed", response_class=StreamingResponse)
async def download_tender_excel_typed(data_type: str, subdir: str):
    tenders: List[Dict[str, Any]] = []
    if data_type == "rot": base_data_dir_for_download = REG_FILTERED_RESULTS_DIR_ROT
    elif data_type == "regular": base_data_dir_for_download = REG_FILTERED_RESULTS_DIR
    else: raise HTTPException(400, "Invalid data type.")
    try: subdir_path = _validate_subdir(subdir, base_dir=base_data_dir_for_download); file_path = subdir_path / FILTERED_TENDERS_FILENAME
    except HTTPException as e: raise e
    except Exception as e: logger.error(f"Path err download {subdir} type {data_type}: {e}"); raise HTTPException(500,"Path error")
    if not subdir_path.is_dir() or not file_path.is_file(): raise HTTPException(404, "Data not found.")
    try:
        with open(file_path, "r", encoding="utf-8") as f: tenders_data = json.load(f)
        if not isinstance(tenders_data, list): raise ValueError("Invalid data format.")
        tenders = tenders_data
    except (json.JSONDecodeError, ValueError) as e: logger.error(f"Err reading {file_path} for download: {e}"); raise HTTPException(500, "Err reading data.")
    except Exception as e: logger.error(f"Err download prep {subdir} type {data_type}: {e}", exc_info=True); raise HTTPException(500, "Err prep download.")
    wb = Workbook(); ws = wb.active; safe_sheet_title = re.sub(r'[\\/*?:\[\]]+', '_', subdir.replace('_Results','').replace('_Tenders', ''))[:31]; ws.title = safe_sheet_title if safe_sheet_title else "Tenders"
    if data_type == "rot":
        headers = ["rot_s_no", "rot_tender_id", "rot_title_ref", "rot_organisation_chain", "rot_tender_stage", "source_site_key", "rot_status_detail_page_link", "rot_stage_summary_file_status", "rot_stage_summary_file_path", "rot_stage_summary_filename"]
    else: 
        headers = ["primary_tender_id", "tender_reference_number", "primary_title", "work_description", "organisation_chain", "source_site_key", "tender_type", "tender_category", "form_of_contract", "contract_type", "tender_value", "emd_amount", "tender_fee", "processing_fee", "primary_published_date", "primary_closing_date", "primary_opening_date", "bid_validity_days", "period_of_work_days", "location", "pincode", "nda_pre_qualification", "independent_external_monitor_remarks", "payment_instruments", "covers_info", "tender_documents", "tender_inviting_authority_name", "tender_inviting_authority_address", "detail_page_link"]
    ws.append(headers)
    for tender in tenders:
        row_data = [];
        if not isinstance(tender, dict): continue
        for header in headers:
            value = tender.get(header, "N/A")
            if data_type == "regular" and isinstance(value, list):
                try:
                    if header == "payment_instruments" and value: value = "; ".join([f"{i.get('s_no', '?')}:{i.get('instrument_type', 'N/A')}" for i in value if isinstance(i,dict)])
                    elif header == "covers_info" and value: value = "; ".join([f"{i.get('cover_no','?')}:{i.get('cover_type','N/A')}" for i in value if isinstance(i,dict)])
                    elif header == "tender_documents" and value: value = "; ".join([f"{d.get('name',d.get('link','Doc'))}({d.get('size','?')})" for d in value if isinstance(d, dict)])
                    else: value = ", ".join(map(str, value))
                    if not value: value = "N/A"
                except Exception: value = "[Error Formatting List]"
            elif data_type == "regular" and isinstance(value, dict):
                try: value = json.dumps(value, ensure_ascii=False)
                except Exception: value = "[Error Formatting Dict]"
            if value is None: value = "N/A"
            elif not isinstance(value, (str, int, float, datetime.datetime, datetime.date)): value = str(value)
            row_data.append(value)
        if len(row_data)==len(headers): ws.append(row_data)
        else: logger.warning(f"Col count mismatch ID {tender.get('primary_tender_id' if data_type=='regular' else 'rot_tender_id','??')} in {subdir} (Type: {data_type}).")
    excel_buffer = io.BytesIO(); wb.save(excel_buffer); excel_buffer.seek(0); 
    safe_subdir_name = re.sub(r'[^\w\-]+', '_', subdir); filename_prefix = "ROT_" if data_type == "rot" else ""; filename = f"{filename_prefix}{safe_subdir_name}_Tenders_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

@app.get("/download/{data_type}/{subdir}", name="download_tender_excel_typed", response_class=StreamingResponse)
async def download_tender_excel_typed(data_type: str, subdir: str):
    tenders: List[Dict[str, Any]] = []
    if data_type == "rot": base_data_dir_for_download = REG_FILTERED_RESULTS_DIR_ROT
    elif data_type == "regular": base_data_dir_for_download = REG_FILTERED_RESULTS_DIR
    else: raise HTTPException(400, "Invalid data type.")
    try: subdir_path = _validate_subdir(subdir, base_dir=base_data_dir_for_download); file_path = subdir_path / FILTERED_TENDERS_FILENAME
    except HTTPException as e: raise e
    except Exception as e: logger.error(f"Path err download {subdir} type {data_type}: {e}"); raise HTTPException(500,"Path error")
    if not subdir_path.is_dir() or not file_path.is_file(): raise HTTPException(404, "Data not found.")
    try:
        with open(file_path, "r", encoding="utf-8") as f: tenders_data = json.load(f)
        if not isinstance(tenders_data, list): raise ValueError("Invalid data format.")
        tenders = tenders_data
    except (json.JSONDecodeError, ValueError) as e: logger.error(f"Err reading {file_path} for download: {e}"); raise HTTPException(500, "Err reading data.")
    except Exception as e: logger.error(f"Err download prep {subdir} type {data_type}: {e}", exc_info=True); raise HTTPException(500, "Err prep download.")
    wb = Workbook(); ws = wb.active; safe_sheet_title = re.sub(r'[\\/*?:\[\]]+', '_', subdir.replace('_Results','').replace('_Tenders', ''))[:31]; ws.title = safe_sheet_title if safe_sheet_title else "Tenders"
    if data_type == "rot":
        headers = ["rot_s_no", "rot_tender_id", "rot_title_ref", "rot_organisation_chain", "rot_tender_stage", "source_site_key", "rot_status_detail_page_link", "rot_stage_summary_file_status", "rot_stage_summary_file_path", "rot_stage_summary_filename"]
    else: 
        headers = ["primary_tender_id", "tender_reference_number", "primary_title", "work_description", "organisation_chain", "source_site_key", "tender_type", "tender_category", "form_of_contract", "contract_type", "tender_value", "emd_amount", "tender_fee", "processing_fee", "primary_published_date", "primary_closing_date", "primary_opening_date", "bid_validity_days", "period_of_work_days", "location", "pincode", "nda_pre_qualification", "independent_external_monitor_remarks", "payment_instruments", "covers_info", "tender_documents", "tender_inviting_authority_name", "tender_inviting_authority_address", "detail_page_link"]
    ws.append(headers)
    for tender in tenders:
        row_data = [];
        if not isinstance(tender, dict): continue
        for header in headers:
            value = tender.get(header, "N/A")
            if data_type == "regular" and isinstance(value, list):
                try:
                    if header == "payment_instruments" and value: value = "; ".join([f"{i.get('s_no', '?')}:{i.get('instrument_type', 'N/A')}" for i in value if isinstance(i,dict)])
                    elif header == "covers_info" and value: value = "; ".join([f"{i.get('cover_no','?')}:{i.get('cover_type','N/A')}" for i in value if isinstance(i,dict)])
                    elif header == "tender_documents" and value: value = "; ".join([f"{d.get('name',d.get('link','Doc'))}({d.get('size','?')})" for d in value if isinstance(d, dict)])
                    else: value = ", ".join(map(str, value))
                    if not value: value = "N/A"
                except Exception: value = "[Error Formatting List]"
            elif data_type == "regular" and isinstance(value, dict):
                try: value = json.dumps(value, ensure_ascii=False)
                except Exception: value = "[Error Formatting Dict]"
            if value is None: value = "N/A"
            elif not isinstance(value, (str, int, float, datetime.datetime, datetime.date)): value = str(value)
            row_data.append(value)
        if len(row_data)==len(headers): ws.append(row_data)
        else: logger.warning(f"Col count mismatch ID {tender.get('primary_tender_id' if data_type=='regular' else 'rot_tender_id','??')} in {subdir} (Type: {data_type}).")
    excel_buffer = io.BytesIO(); wb.save(excel_buffer); excel_buffer.seek(0); 
    safe_subdir_name = re.sub(r'[^\w\-]+', '_', subdir); filename_prefix = "ROT_" if data_type == "rot" else ""; filename = f"{filename_prefix}{safe_subdir_name}_Tenders_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

# In dashboard.py

@app.post("/bulk-delete-typed", name="bulk_delete_tender_sets_typed")
async def bulk_delete_tender_sets_typed(request: Request, data_type: str = Form(...), selected_subdirs: List[str] = Form(...)):
    if not selected_subdirs:
        raise HTTPException(status_code=400, detail="No filter sets selected.")

    logger.info(f"Bulk delete request for {data_type.upper()} sets: {selected_subdirs}")
    base_data_dir_for_delete: Path # Define type

    if data_type == "rot":
        base_data_dir_for_delete = ROT_FILTERED_RESULTS_DIR # MODIFIED
    elif data_type == "regular":
        base_data_dir_for_delete = REG_FILTERED_RESULTS_DIR # MODIFIED
    else:
        # Ensure status_code for HTTPException
        raise HTTPException(status_code=400, detail="Invalid data type for bulk delete.")

    deleted_count_bulk = 0
    errors_delete: List[str] = [] # Define type for clarity

    for subdir_name_del in selected_subdirs:
        try:
            subdir_path_to_del = _validate_subdir(subdir_name_del, base_dir=base_data_dir_for_delete)
            if subdir_path_to_del.is_dir():
                shutil.rmtree(subdir_path_to_del)
                logger.info(f"Deleted {data_type.upper()} set: {subdir_path_to_del}")
                deleted_count_bulk += 1
            else:
                logger.warning(f"Directory not found for {data_type.upper()} delete: {subdir_path_to_del} (selected name: {subdir_name_del})")
                errors_delete.append(f"Not found: {subdir_name_del}")
        except HTTPException as http_e_del: # Catch validation errors from _validate_subdir
            logger.error(f"Validation error during bulk delete of {data_type.upper()} set '{subdir_name_del}': {http_e_del.detail}")
            errors_delete.append(f"Invalid name: {subdir_name_del}")
        except Exception as e_del: # Catch other errors like permission issues during rmtree
            logger.error(f"Error deleting {data_type.upper()} set '{subdir_name_del}' at path '{subdir_path_to_del if 'subdir_path_to_del' in locals() else 'unknown'}': {e_del}", exc_info=True)
            errors_delete.append(f"Error deleting: {subdir_name_del}")

    redirect_url_path_del = app.url_path_for("homepage")
    query_params_del: Dict[str, str] = {} # Define type
    msg_type_key = f"bulk_delete_status_{data_type}" # e.g., bulk_delete_status_rot

    if errors_delete:
        query_params_del[msg_type_key] = "partial"
        query_params_del["errors"] = str(len(errors_delete))
        query_params_del["deleted"] = str(deleted_count_bulk)
        # Optionally include details of errors if desired, but can make URL long
        # query_params_del["error_details"] = ";".join(errors_delete)[:200] # Example
    elif deleted_count_bulk > 0:
        query_params_del[msg_type_key] = "ok"
        query_params_del["deleted"] = str(deleted_count_bulk)
    else: # No errors, but nothing deleted (e.g., selected items were already gone)
        query_params_del[msg_type_key] = "none_deleted"
        # No need to set 'deleted' or 'errors' if both are 0

    redirect_url_del = URL(redirect_url_path_del).replace_query_params(**query_params_del)
    return RedirectResponse(url=str(redirect_url_del), status_code=status.HTTP_303_SEE_OTHER)

@app.get("/download/{data_type}/{subdir}", name="download_tender_excel_typed", response_class=StreamingResponse)
async def download_tender_excel_typed(data_type: str, subdir: str):
    tenders: List[Dict[str, Any]] = []
    base_data_dir_for_download: Path # Define type

    if data_type == "rot":
        base_data_dir_for_download = ROT_FILTERED_RESULTS_DIR # MODIFIED
    elif data_type == "regular":
        base_data_dir_for_download = REG_FILTERED_RESULTS_DIR # MODIFIED
    else:
        raise HTTPException(status_code=400, detail="Invalid data type for download.") # Added status_code

    try:
        subdir_path = _validate_subdir(subdir, base_dir=base_data_dir_for_download)
        file_path = subdir_path / FILTERED_TENDERS_FILENAME
    except HTTPException as e_val: # Catch validation specific errors
        logger.warning(f"Validation error for excel download '{subdir}' type '{data_type}': {e_val.detail}")
        raise e_val
    except Exception as e_path:
        logger.error(f"Path error for excel download '{subdir}' type '{data_type}': {e_path}", exc_info=True)
        raise HTTPException(status_code=500,detail="Path error for excel download.")

    if not subdir_path.is_dir() or not file_path.is_file():
        logger.warning(f"Data not found for excel download: {file_path}")
        raise HTTPException(status_code=404, detail="Data not found for excel generation.")

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            tenders_data = json.load(f)
        if not isinstance(tenders_data, list):
            logger.error(f"Invalid data format (not list) in {file_path} for excel download.")
            raise ValueError("Invalid data format.")
        tenders = tenders_data
    except (json.JSONDecodeError, ValueError) as e_parse:
        logger.error(f"Error reading/parsing {file_path} for excel download: {e_parse}")
        raise HTTPException(status_code=500, detail="Error reading data for excel generation.")
    except Exception as e_load:
        logger.error(f"Error during excel download prep for '{subdir}' (Type: {data_type}) from {file_path}: {e_load}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error preparing download.")

    wb = Workbook()
    ws = wb.active
    safe_sheet_title = re.sub(r'[\\/*?:\[\]]+', '_', subdir.replace('_Results','').replace('_Tenders', ''))[:31]
    ws.title = safe_sheet_title if safe_sheet_title else "Tenders"

    headers: List[str] # Define type
    if data_type == "rot":
        headers = [
            "rot_s_no", "rot_tender_id", "rot_title_ref", "rot_organisation_chain",
            "rot_tender_stage", "source_site_key", "rot_status_detail_page_link",
            "rot_stage_summary_file_status", "rot_stage_summary_file_path", "rot_stage_summary_filename"
        ]
    else: # regular
        # ENSURE THIS LIST IS COMPLETE AND MATCHES YOUR ACTUAL INTENDED HEADERS
        headers = [
            "primary_tender_id", "tender_reference_number", "primary_title", "work_description",
            "organisation_chain", "source_site_key", "tender_type", "tender_category",
            "form_of_contract", "contract_type", "tender_value", "emd_amount", "tender_fee",
            "processing_fee", "primary_published_date", "primary_closing_date", "primary_opening_date",
            "bid_validity_days", "period_of_work_days", "location", "pincode", # Assuming pincode was the end
            "nda_pre_qualification", "independent_external_monitor_remarks", # Added based on filter_engine
            "payment_instruments", "covers_info", "tender_documents", # Added based on filter_engine
            "tender_inviting_authority_name", "tender_inviting_authority_address", "detail_page_link" # Added
        ]
    ws.append(headers)

    for tender in tenders:
        row_data = []
        if not isinstance(tender, dict): continue
        for header in headers:
            value = tender.get(header, "N/A")
            # Formatting for list/dict types (primarily for regular tenders)
            if data_type == "regular": # This complex formatting typically applies to regular tenders
                if isinstance(value, list):
                    try:
                        if header == "payment_instruments" and value:
                            value = "; ".join([f"{i.get('s_no', '?')}:{i.get('instrument_type', 'N/A')}" for i in value if isinstance(i,dict)])
                        elif header == "covers_info" and value:
                            value = "; ".join([f"{i.get('cover_no','?')}:{i.get('cover_type','N/A')}" for i in value if isinstance(i,dict)])
                        elif header == "tender_documents" and value:
                            value = "; ".join([f"{d.get('name',d.get('link','Doc'))}({d.get('size','?')})" for d in value if isinstance(d, dict)])
                        else: # Generic list to string
                            value = ", ".join(map(str, value))
                        if not value: value = "N/A" # Ensure empty lists become "N/A"
                    except Exception: value = "[Error Formatting List]"
                elif isinstance(value, dict):
                    try: value = json.dumps(value, ensure_ascii=False)
                    except Exception: value = "[Error Formatting Dict]"
            
            # General value cleaning for all types
            if value is None: value = "N/A"
            elif not isinstance(value, (str, int, float, datetime.datetime, datetime.date)): # Check for datetime too
                value = str(value)
            row_data.append(value)

        if len(row_data) == len(headers):
            ws.append(row_data)
        else:
            tender_id_for_log = tender.get('primary_tender_id' if data_type=='regular' else 'rot_tender_id', 'UnknownID')
            logger.warning(f"Column count mismatch for tender ID '{tender_id_for_log}' in excel export for '{subdir}' (Type: {data_type}). Expected {len(headers)}, got {len(row_data)}.")

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    safe_subdir_name = re.sub(r'[^\w\-]+', '_', subdir)
    filename_prefix = "ROT_" if data_type == "rot" else ""
    filename = f"{filename_prefix}{safe_subdir_name}_Tenders_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.post("/bulk-delete-typed", name="bulk_delete_tender_sets_typed") 
async def bulk_delete_tender_sets_typed(request: Request, data_type: str = Form(...), selected_subdirs: List[str] = Form(...)):
    if not selected_subdirs: raise HTTPException(status_code=400, detail="No filter sets selected.")
    logger.info(f"Bulk delete request for {data_type.upper()} sets: {selected_subdirs}")
    if data_type == "rot": base_data_dir_for_delete = REG_FILTERED_RESULTS_DIR_ROT
    elif data_type == "regular": base_data_dir_for_delete = REG_FILTERED_RESULTS_DIR
    else: raise HTTPException(400, "Invalid data type for bulk delete.")
    deleted_count_bulk = 0; errors_delete = []
    for subdir_name_del in selected_subdirs:
        try:
            subdir_path_to_del = _validate_subdir(subdir_name_del, base_dir=base_data_dir_for_delete)
            if subdir_path_to_del.is_dir(): shutil.rmtree(subdir_path_to_del); logger.info(f"Deleted {data_type.upper()} set: {subdir_path_to_del}"); deleted_count_bulk += 1
            else: logger.warning(f"Dir not found for {data_type.upper()} delete: {subdir_path_to_del} (selected: {subdir_name_del})"); errors_delete.append(f"Not found: {subdir_name_del}")
        except HTTPException as http_e_del: logger.error(f"Validation err bulk delete {data_type.upper()} '{subdir_name_del}': {http_e_del.detail}"); errors_delete.append(f"Invalid: {subdir_name_del}")
        except Exception as e_del: logger.error(f"Error deleting {data_type.upper()} set '{subdir_name_del}': {e_del}", exc_info=True); errors_delete.append(f"Error: {subdir_name_del}")
    redirect_url_path_del = app.url_path_for("homepage"); query_params_del = {}; msg_type_key = f"bulk_delete_status_{data_type}"
    if errors_delete: query_params_del[msg_type_key] = "partial"; query_params_del["errors"] = str(len(errors_delete)); query_params_del["deleted"] = str(deleted_count_bulk)
    elif deleted_count_bulk > 0: query_params_del[msg_type_key] = "ok"; query_params_del["deleted"] = str(deleted_count_bulk)
    else: query_params_del[msg_type_key] = "none_deleted"
    redirect_url_del = URL(redirect_url_path_del).replace_query_params(**query_params_del)
    return RedirectResponse(url=str(redirect_url_del), status_code=status.HTTP_303_SEE_OTHER)

@app.post("/bulk-download-typed", name="bulk_download_tender_excel_typed", response_class=StreamingResponse)
async def bulk_download_tender_excel_typed(request: Request, data_type: str = Form(...), selected_subdirs: List[str] = Form(...)):
    if not selected_subdirs:
        raise HTTPException(status_code=400, detail="No filter sets selected.") # Added status_code

    logger.info(f"Bulk download requested for {data_type.upper()} sets: {selected_subdirs}")
    base_data_dir_for_bulk: Path # Define type

    if data_type == "rot":
        base_data_dir_for_bulk = ROT_FILTERED_RESULTS_DIR # MODIFIED
    elif data_type == "regular":
        base_data_dir_for_bulk = REG_FILTERED_RESULTS_DIR # MODIFIED
    else:
        # Ensure status_code for HTTPException
        raise HTTPException(status_code=400, detail="Invalid data type for bulk download.")

    wb = Workbook()
    wb.remove(wb.active) # Start with a clean workbook
    processed_sheets = 0
    errors_bulk: List[str] = [] # Define type

    headers: List[str] # Define type
    if data_type == "rot":
        headers = [
            "rot_s_no", "rot_tender_id", "rot_title_ref", "rot_organisation_chain",
            "rot_tender_stage", "source_site_key", "rot_status_detail_page_link",
            "rot_stage_summary_file_status", "rot_stage_summary_file_path", "rot_stage_summary_filename"
        ]
    else: # regular
        # ENSURE THIS LIST IS COMPLETE AND MATCHES YOUR ACTUAL INTENDED HEADERS (same as single download)
        headers = [
            "primary_tender_id", "tender_reference_number", "primary_title", "work_description",
            "organisation_chain", "source_site_key", "tender_type", "tender_category",
            "form_of_contract", "contract_type", "tender_value", "emd_amount", "tender_fee",
            "processing_fee", "primary_published_date", "primary_closing_date", "primary_opening_date",
            "bid_validity_days", "period_of_work_days", "location", "pincode",
            "nda_pre_qualification", "independent_external_monitor_remarks",
            "payment_instruments", "covers_info", "tender_documents",
            "tender_inviting_authority_name", "tender_inviting_authority_address", "detail_page_link"
        ]

    for subdir_name in selected_subdirs:
        tenders_list_sheet: List[Dict[str, Any]] = [] # Define type
        try:
            subdir_path_validated = _validate_subdir(subdir_name, base_dir=base_data_dir_for_bulk)
            file_path_to_read = subdir_path_validated / FILTERED_TENDERS_FILENAME

            if not subdir_path_validated.is_dir():
                logger.warning(f"Bulk download: Directory missing for '{subdir_name}' at {subdir_path_validated}")
                errors_bulk.append(f"Missing Dir: '{subdir_name}'")
                continue
            if not file_path_to_read.is_file():
                logger.warning(f"Bulk download: Data file missing for '{subdir_name}' at {file_path_to_read}")
                errors_bulk.append(f"Data Missing: '{subdir_name}'")
                continue

            with open(file_path_to_read, "r", encoding="utf-8") as f_sheet:
                tenders_data_sheet = json.load(f_sheet)

            if not isinstance(tenders_data_sheet, list):
                logger.warning(f"Bulk download: Invalid data format in '{file_path_to_read}' for '{subdir_name}'")
                errors_bulk.append(f"Invalid Data: '{subdir_name}'")
                continue

            tenders_list_sheet = tenders_data_sheet
            safe_sheet_title_bulk = re.sub(r'[\\/*?:\[\]]+', '_', subdir_name.replace('_Results','').replace('_Tenders',''))[:31]
            ws_bulk = wb.create_sheet(title=safe_sheet_title_bulk if safe_sheet_title_bulk else f"Sheet_{processed_sheets+1}")
            ws_bulk.append(headers)
            rows_added_sheet = 0

            for tender_item_sheet in tenders_list_sheet:
                  row_data_sheet = []
                  if not isinstance(tender_item_sheet, dict): continue
                  for header_name in headers:
                     value_cell = tender_item_sheet.get(header_name, "N/A")
                     # Formatting for list/dict types (primarily for regular tenders)
                     if data_type == "regular":
                        if isinstance(value_cell, list):
                            try:
                                if header_name == "payment_instruments" and value_cell:
                                    value_cell = "; ".join([f"{i.get('s_no', '?')}:{i.get('instrument_type', 'N/A')}" for i in value_cell if isinstance(i,dict)])
                                elif header_name == "covers_info" and value_cell:
                                    value_cell = "; ".join([f"{i.get('cover_no','?')}:{i.get('cover_type','N/A')}" for i in value_cell if isinstance(i,dict)])
                                elif header_name == "tender_documents" and value_cell:
                                    value_cell = "; ".join([f"{d.get('name', d.get('link', 'Doc'))} ({d.get('size','?')})" for d in value_cell if isinstance(d, dict)])
                                else: # Generic list to string
                                    value_cell = ", ".join(map(str, value_cell))
                                if not value_cell: value_cell = "N/A"
                            except Exception: value_cell = "[Error Formatting List]"
                        elif isinstance(value_cell, dict):
                            try: value_cell = json.dumps(value_cell, ensure_ascii=False)
                            except Exception: value_cell = "[Error Formatting Dict]"

                     if value_cell is None: value_cell = "N/A"
                     elif not isinstance(value_cell, (str, int, float, datetime.datetime, datetime.date)):
                         value_cell = str(value_cell)
                     row_data_sheet.append(value_cell)

                  if len(row_data_sheet) == len(headers):
                      ws_bulk.append(row_data_sheet)
                      rows_added_sheet += 1
                  else:
                      tender_id_for_log_bulk = tender_item_sheet.get('primary_tender_id' if data_type=='regular' else 'rot_tender_id', 'UnknownID')
                      logger.warning(f"Column count mismatch in bulk excel for tender ID '{tender_id_for_log_bulk}', sheet '{safe_sheet_title_bulk}'. Expected {len(headers)}, got {len(row_data_sheet)}.")
            
            logger.info(f"Added sheet '{safe_sheet_title_bulk}' ({rows_added_sheet} {data_type} items) to bulk download.")
            processed_sheets += 1
        except HTTPException as http_e_bulk:
            logger.error(f"Bulk download validation error for '{subdir_name}': {http_e_bulk.detail}")
            errors_bulk.append(f"Invalid: {subdir_name}")
        except Exception as e_bulk:
            logger.error(f"Error processing sheet for '{subdir_name}' in bulk download: {e_bulk}", exc_info=True)
            errors_bulk.append(f"Error processing: {subdir_name}")

    if processed_sheets == 0:
        err_details_bulk = f"Bulk download failed for {data_type.upper()}. No valid filter sets could be processed." + (f" Issues encountered: {'; '.join(errors_bulk)}" if errors_bulk else "")
        logger.error(err_details_bulk)
        # Ensure status_code for HTTPException
        raise HTTPException(status_code=400, detail=err_details_bulk)

    if errors_bulk:
        logger.warning(f"Bulk download for {data_type.upper()} completed with some issues: {'; '.join(errors_bulk)}")

    excel_buffer_bulk = io.BytesIO()
    wb.save(excel_buffer_bulk)
    excel_buffer_bulk.seek(0)
    
    timestamp_bulk = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    filename_bulk = f"Bulk_{data_type.upper()}_Download_{timestamp_bulk}.xlsx"
    logger.info(f"Sending bulk {data_type.upper()} download: {filename_bulk} ({processed_sheets} sheets)")
    
    return StreamingResponse(
        excel_buffer_bulk,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename_bulk}\""}
    )

@app.post("/delete/{data_type}/{subdir}", name="delete_tender_set_typed")
async def delete_tender_set_typed(request: Request, data_type: str, subdir: str):
    base_data_dir_single_del: Path

    if data_type == "rot":
        base_data_dir_single_del = ROT_FILTERED_RESULTS_DIR
    elif data_type == "regular":
        base_data_dir_single_del = REG_FILTERED_RESULTS_DIR
    else:
        raise HTTPException(status_code=400, detail="Invalid data type for delete operation.")

    subdir_path_single_del: Optional[Path] = None # Initialize for use in except block
    try:
        subdir_path_single_del = _validate_subdir(subdir, base_dir=base_data_dir_single_del)
        if subdir_path_single_del.is_dir():
            shutil.rmtree(subdir_path_single_del)
            logger.info(f"Deleted {data_type.upper()} filter set: {subdir_path_single_del}")
            
            # MODIFIED: Use _clean_path_component for the message
            cleaned_subdir_for_msg = _clean_path_component(subdir)
            redirect_url = URL(app.url_path_for('homepage')).include_query_params(
                msg=f"filter_set_{data_type}_{cleaned_subdir_for_msg}_deleted_successfully"
            )
            return RedirectResponse(url=str(redirect_url), status_code=status.HTTP_303_SEE_OTHER)
        else:
            logger.warning(f"{data_type.upper()} filter set directory not found for deletion: {subdir_path_single_del}")
            raise HTTPException(status_code=404, detail=f"{data_type.upper()} filter set '{subdir}' not found.")
    except HTTPException as e_single_del_http:
        logger.warning(f"HTTPException during delete of '{subdir}' (Type: {data_type}): {e_single_del_http.detail}")
        raise e_single_del_http
    except Exception as e_single_del:
        path_info = str(subdir_path_single_del) if subdir_path_single_del else "unknown_path (validation may have failed)"
        logger.error(f"Error deleting {data_type.upper()} set '{subdir}' at path '{path_info}': {e_single_del}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Could not delete {data_type.upper()} filter set '{subdir}'.")

@app.get("/run-filter", name="run_filter_form", response_class=HTMLResponse)
async def run_filter_form(request: Request):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine error.")
    if not filter_engine_available:
        return templates.TemplateResponse("error.html", {"request": request, "error": "Filter engine components are not available. Cannot create new filters."}, status_code=503)

    settings = load_settings()
    site_configs_dict = settings.get('site_configurations', {})
    # This 'available_sites' is for the "Filter by Source Site (In-Data Filter)" dropdown
    available_sites_for_in_data_filter = sorted(list(site_configs_dict.keys())) if site_configs_dict and isinstance(site_configs_dict, dict) else []

    # --- Regular Tender Source File Check ---
    no_source_file_regular = True
    latest_source_filename_regular = None
    if REG_FINAL_GLOBAL_MERGED_DIR.is_dir(): # Checks site_data/REG/FinalGlobalMerged/
        try:
            source_files_reg = sorted(
                [p for p in REG_FINAL_GLOBAL_MERGED_DIR.glob("Final_Tender_List_*.txt") if p.is_file()],
                key=lambda p: p.stat().st_mtime, reverse=True
            )
            if source_files_reg:
                latest_source_filename_regular = source_files_reg[0].name
                no_source_file_regular = False
        except OSError as e: logger.error(f"Error listing regular source files from {REG_FINAL_GLOBAL_MERGED_DIR}: {e}")
    else: logger.warning(f"Regular global merged directory not found: '{REG_FINAL_GLOBAL_MERGED_DIR}'")

    # --- ROT Tender Source File Check (ONLY Global File) ---
    no_source_file_rot = True
    latest_source_filename_rot = None
    if ROT_FINAL_GLOBAL_MERGED_DIR.is_dir(): # Checks site_data/ROT/FinalGlobalMerged/
        try:
            global_rot_files = sorted(
                [p for p in ROT_FINAL_GLOBAL_MERGED_DIR.glob("Final_ROT_Tender_List_*.txt") if p.is_file()],
                key=lambda p: p.stat().st_mtime, reverse=True
            )
            if global_rot_files:
                latest_source_filename_rot = global_rot_files[0].name
                no_source_file_rot = False
        except OSError as e: logger.error(f"Error listing global ROT source files from {ROT_FINAL_GLOBAL_MERGED_DIR}: {e}")
    else: logger.warning(f"Global ROT merged directory not found: '{ROT_FINAL_GLOBAL_MERGED_DIR}'")
    
    return templates.TemplateResponse(
        "run_filter.html",
        {
            "request": request,
            "available_sites": available_sites_for_in_data_filter, # For the "Filter by Source Site" dropdown
            "no_source_file_regular": no_source_file_regular,
            "latest_source_filename_regular": latest_source_filename_regular,
            "no_source_file_rot": no_source_file_rot, 
            "latest_source_filename_rot": latest_source_filename_rot
        }
    )
    
    # This flag is true if NO site-specific ROT files are available for filtering
    no_source_file_rot = not available_site_specific_rot_files

    return templates.TemplateResponse(
        "run_filter.html",
        {
            "request": request,
            "available_sites_for_site_filter_dropdown": available_sites_for_site_filter_dropdown,
            "no_source_file_regular": no_source_file_regular,
            "latest_source_filename_regular": latest_source_filename_regular,
            "available_site_specific_rot_files": available_site_specific_rot_files, # List of {display_name, filename}
            "no_source_file_rot": no_source_file_rot # True if available_site_specific_rot_files is empty
        }
    )

@app.post("/run-filter", name="run_filter_submit", response_class=HTMLResponse)
async def run_filter_submit(
    request: Request,
    data_type: str = Form(...),
    # Note: rot_source_file or rot_source_file_site_specific parameter is REMOVED
    keywords: str = Form(""),
    regex: bool = Form(False),
    filter_name: str = Form(...),
    site_key: Optional[str] = Form(None), # This is for the in-data <SourceSiteKey> filter
    start_date: Optional[str] = Form(None),
    end_date: Optional[str] = Form(None)
):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine error.")
    if not filter_engine_available: # filter_engine_available should be defined globally
        return templates.TemplateResponse("error.html", {"request": request, "error": "Filter engine components are not available. Cannot run filter."}, status_code=503)

    # Validate filter_name (this existing validation is good)
    if not filter_name or re.search(r'[<>:"/\\|?*]', filter_name) or ".." in filter_name or filter_name.strip() != filter_name:
        logger.warning(f"Invalid filter name submitted: '{filter_name}'")
        return templates.TemplateResponse("error.html", {"request": request, "error": "Invalid Filter Name. Avoid special characters like <>:\"/\\|?* and leading/trailing spaces."}, status_code=400)

    if data_type not in ["regular", "rot"]:
        logger.warning(f"Invalid data_type submitted for filter: '{data_type}'")
        return templates.TemplateResponse("error.html", {"request": request, "error": "Invalid data type specified for filter."}, status_code=400)

    base_folder_for_filter_input: Path
    source_file_pattern: str
    
    if data_type == "rot":
        base_folder_for_filter_input = ROT_FINAL_GLOBAL_MERGED_DIR # Always use global merged for ROT
        source_file_pattern = "Final_ROT_Tender_List_*.txt"
    else: # regular
        base_folder_for_filter_input = REG_FINAL_GLOBAL_MERGED_DIR # Always use global merged for Regular
        source_file_pattern = "Final_Tender_List_*.txt"

    latest_source_file_path_actual: Optional[Path] = None

    if base_folder_for_filter_input.is_dir():
        try:
            source_files_list = sorted(
                [p for p in base_folder_for_filter_input.glob(source_file_pattern) if p.is_file()],
                key=lambda p: p.stat().st_mtime, reverse=True
            )
            if source_files_list:
                latest_source_file_path_actual = source_files_list[0]
            else:
                error_message = f"No consolidated source file ('{source_file_pattern}') found in {base_folder_for_filter_input} for {data_type} tenders. "
                if data_type == "rot":
                    error_message += "Please run 'Prepare/Refresh ROT Data for AI' on the AI Analysis page first to create this file."
                else: # regular
                    error_message += "Please ensure regular scrapers have run and data has been globally merged by the site controller."
                logger.error(error_message)
                return templates.TemplateResponse("error.html", {"request": request, "error": error_message}, status_code=404)
        except Exception as e_find_source:
            logger.error(f"Error finding {data_type} source file in {base_folder_for_filter_input}: {e_find_source}", exc_info=True)
            return templates.TemplateResponse("error.html", {"request": request, "error": f"Error accessing {data_type} source data."}, status_code=500)
    else:
        logger.error(f"{data_type.capitalize()} source data directory for filters not found: {base_folder_for_filter_input}")
        return templates.TemplateResponse("error.html", {"request": request, "error": f"{data_type.capitalize()} source data directory for filters is missing."}, status_code=500)

    # This check ensures that latest_source_file_path_actual was successfully found
    if not latest_source_file_path_actual or not latest_source_file_path_actual.is_file():
        logger.error(f"Source file for filtering (Type: {data_type}) could not be determined or does not exist. Path checked: {latest_source_file_path_actual}")
        return templates.TemplateResponse("error.html", {"request": request, "error": "Selected source data file for filtering is missing or invalid."}, status_code=404)

    try:
        keyword_list_cleaned = [kw.strip() for kw in keywords.split(",") if kw.strip()]
        
        # Call filter_engine.run_filter with the direct path to the source file
        # filter_engine.run_filter was already modified to accept tender_file_to_filter
        result_path_str_actual = run_filter(
            tender_file_to_filter=latest_source_file_path_actual,
            keywords=keyword_list_cleaned,
            use_regex=regex,
            filter_name=filter_name,
            site_key=site_key or None,
            start_date=start_date or None,
            end_date=end_date or None,
            data_type=data_type
        )
        
        if not result_path_str_actual or not Path(result_path_str_actual).is_file():
            logger.error(f"Filter engine run for '{filter_name}' ({data_type.upper()}) did not produce a valid output file path. Received: {result_path_str_actual}")
            raise RuntimeError(f"Filter engine failed for {data_type} data or did not return a valid file path.")

        logger.info(f"Filter '{filter_name}' ({data_type.upper()}) using source '{latest_source_file_path_actual.name}' completed. Output: {result_path_str_actual}")
        
        created_subdir_name = Path(result_path_str_actual).parent.name
        return templates.TemplateResponse("success.html", {"request": request, "subdir": created_subdir_name, "data_type": data_type})

    except FileNotFoundError as e_fnf:
        logger.error(f"Filter run error for '{filter_name}': Source file specified for filter_engine not found. Detail: {e_fnf}", exc_info=True)
        return templates.TemplateResponse("error.html", {"request": request, "error": f"Filter source file error: {e_fnf}. This may indicate an issue with how the source file path was determined."}, status_code=404)
    except Exception as e_run_filt:
        source_filename_for_log = latest_source_file_path_actual.name if latest_source_file_path_actual else "unknown_source_file"
        logger.error(f"Filter run error for '{filter_name}' ({data_type.upper()}) using source '{source_filename_for_log}': {e_run_filt}", exc_info=True)
        return templates.TemplateResponse("error.html", {"request": request, "error": f"Filter run failed: {type(e_run_filt).__name__}. Check server logs for details."}, status_code=500)

@app.get("/regex-help", name="regex_help_page", response_class=HTMLResponse) 
async def regex_help_page(request: Request):
    if not templates: raise HTTPException(503, "Template engine error.")
    return templates.TemplateResponse("regex_help.html", {"request": request})

@app.post("/save-retention-settings", name="save_retention_settings") 
async def save_retention_settings(request: Request, enable_retention: bool = Form(False), retention_days: int = Form(30)):
    if retention_days < 1: logger.warning(f"Invalid retention: {retention_days}. Use 1."); retention_days = 1
    elif retention_days > 3650: logger.warning(f"Excessive retention: {retention_days}. Use 3650."); retention_days = 3650
    current_settings = load_settings()
    if "retention" not in current_settings or not isinstance(current_settings.get("retention"), dict): current_settings["retention"] = {}
    current_settings["retention"]["enabled"] = enable_retention; current_settings["retention"]["days"] = retention_days
    if save_settings(current_settings): return RedirectResponse(url=app.url_path_for('settings_page') + "?msg=retention_settings_saved", status_code=303)
    else: return RedirectResponse(url=app.url_path_for('settings_page') + "?error=retention_settings_save_failed", status_code=303)

@app.post("/save-scraper-parameters", name="save_scraper_parameters") 
async def save_scraper_parameters(
    request: Request, concurrency_list: int = Form(...), concurrency_detail: int = Form(...), 
    max_list_pages: int = Form(...), retries: int = Form(...), 
    page_load_timeout: int = Form(...), detail_page_timeout: int = Form(...), 
    rot_max_list_pages: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_scrape_limits"]["max_list_pages"]), 
    rot_retries: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_scrape_limits"]["retries"]), 
    rot_detail_processing: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_concurrency"]["detail_processing"]), 
    rot_page_load_timeout: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_timeouts"]["page_load"]), 
    rot_detail_page_timeout: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_timeouts"]["detail_page"]), 
    rot_download_timeout: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_timeouts"]["download"]), 
    rot_pagination_load_wait: int = Form(DEFAULT_SETTINGS["global_scraper_settings"]["rot_timeouts"]["pagination_load_wait"]),
    rot_default_status_value: str = Form(DEFAULT_SETTINGS["scheduler"]["rot_default_status_value"])
):
    cl = max(1, min(concurrency_list, 50)); cd = max(1, min(concurrency_detail, 20)); mlp = max(1, max_list_pages); rt = max(0, min(retries, 5)); plt = max(10000, page_load_timeout); dpt = max(10000, detail_page_timeout)
    if cl!=concurrency_list or cd!=concurrency_detail or mlp!=max_list_pages or rt!=retries or plt!=page_load_timeout or dpt!=detail_page_timeout: logger.warning("Regular scraper params adjusted.")
    rot_mlp = max(1, rot_max_list_pages); rot_rt = max(0, min(rot_retries, 5)); rot_dp = max(1, min(rot_detail_processing, 10)); rot_plt = max(10000, rot_page_load_timeout); rot_dpt = max(10000, rot_detail_page_timeout); rot_dlt = max(30000, rot_download_timeout); rot_plw = max(1000, rot_pagination_load_wait)
    current_settings = load_settings()
    if "global_scraper_settings" not in current_settings or not isinstance(current_settings["global_scraper_settings"], dict): current_settings["global_scraper_settings"] = {}
    gss = current_settings["global_scraper_settings"]
    if "concurrency" not in gss or not isinstance(gss["concurrency"], dict): gss["concurrency"] = {}
    if "scrape_py_limits" not in gss or not isinstance(gss["scrape_py_limits"], dict): gss["scrape_py_limits"] = {}
    if "timeouts" not in gss or not isinstance(gss["timeouts"], dict): gss["timeouts"] = {}
    gss["concurrency"]["list_pages"] = cl; gss["concurrency"]["detail_pages"] = cd; gss["scrape_py_limits"]["max_list_pages"] = mlp; gss["scrape_py_limits"]["retries"] = rt; gss["timeouts"]["page_load"] = plt; gss["timeouts"]["detail_page"] = dpt
    if "rot_scrape_limits" not in gss or not isinstance(gss["rot_scrape_limits"], dict): gss["rot_scrape_limits"] = {}
    if "rot_concurrency" not in gss or not isinstance(gss["rot_concurrency"], dict): gss["rot_concurrency"] = {}
    if "rot_timeouts" not in gss or not isinstance(gss["rot_timeouts"], dict): gss["rot_timeouts"] = {}
    gss["rot_scrape_limits"]["max_list_pages"] = rot_mlp; gss["rot_scrape_limits"]["retries"] = rot_rt; gss["rot_concurrency"]["detail_processing"] = rot_dp; gss["rot_timeouts"]["page_load"] = rot_plt; gss["rot_timeouts"]["detail_page"] = rot_dpt; gss["rot_timeouts"]["download"] = rot_dlt; gss["rot_timeouts"]["pagination_load_wait"] = rot_plw
    if "scheduler" not in current_settings or not isinstance(current_settings["scheduler"], dict): current_settings["scheduler"] = {}
    current_settings["scheduler"]["rot_default_status_value"] = rot_default_status_value
    if save_settings(current_settings): logger.info(f"Global scraper (REG+ROT) params saved."); return RedirectResponse(url=app.url_path_for('settings_page') + "?msg=scraper_params_saved", status_code=303)
    else: return RedirectResponse(url=app.url_path_for('settings_page') + "?error=scraper_params_save_failed", status_code=303)

@app.post("/save-enabled-sites", name="save_enabled_sites") 
async def save_enabled_sites(request: Request):
    form_data = await request.form()
    current_settings = load_settings()
    if "site_configurations" not in current_settings or not isinstance(current_settings["site_configurations"], dict):
        logger.error("Cannot save: 'site_configurations' key is missing or has an invalid type in settings.json.")
        return RedirectResponse(url=app.url_path_for('settings_page') + "?error=site_config_missing_or_invalid", status_code=303)
    
    updated_count = 0
    # Ensure all known site_keys from settings are processed
    for site_key_from_settings in current_settings["site_configurations"]:
        # Check if the checkbox for this site_key was submitted in the form
        # HTML checkboxes only submit a value if they are checked.
        is_enabled_in_form = form_data.get(f"site_{site_key_from_settings}") == "on"
        
        # Update the 'enabled' status in the current_settings dictionary
        if site_key_from_settings in current_settings["site_configurations"]: # Should always be true here
            current_settings["site_configurations"][site_key_from_settings]["enabled"] = is_enabled_in_form
            updated_count += 1
        else:
            # This case should ideally not happen if looping through keys from current_settings
            logger.warning(f"Site key '{site_key_from_settings}' found in settings but not processed for enabling/disabling. This is unexpected.")

    logger.info(f"Enabled sites settings update processed for {updated_count} sites based on current configuration.")
    
    if save_settings(current_settings):
        logger.info("Enabled sites settings were saved successfully.")
        return RedirectResponse(url=app.url_path_for('settings_page') + "?msg=enabled_sites_saved", status_code=303)
    else:
        logger.error("Failed to save the updated enabled sites settings to settings.json.")
        return RedirectResponse(url=app.url_path_for('settings_page') + "?error=enabled_sites_save_failed", status_code=303)

@app.post("/save-main-schedule", name="save_main_schedule") 
async def save_main_schedule(request: Request, freq_main: str = Form(...), time_main: str = Form("02:00"), day_main_weekly: str = Form("7"), day_main_monthly: str = Form("1")):
    form_data = await request.form(); current_settings = load_settings()
    if "scheduler" not in current_settings or not isinstance(current_settings.get("scheduler"), dict): current_settings["scheduler"] = {}
    is_enabled_main = form_data.get("enable_main_schedule") == "True" 
    current_settings["scheduler"]["main_enabled"] = is_enabled_main
    if is_enabled_main:
        valid_freq = ["daily", "weekly", "monthly"];
        if freq_main not in valid_freq: freq_main = "daily"
        try: datetime.datetime.strptime(time_main, "%H:%M")
        except ValueError: time_main = "02:00"
        valid_days_weekly = [str(i) for i in range(1, 8)];
        if day_main_weekly not in valid_days_weekly: day_main_weekly = "7"
        valid_days_monthly = [str(i) for i in range(1, 29)] + ["last"];
        if day_main_monthly not in valid_days_monthly: day_main_monthly = "1"
        current_settings["scheduler"]["main_frequency"] = freq_main
        current_settings["scheduler"]["main_time"] = time_main
        current_settings["scheduler"]["main_day_weekly"] = day_main_weekly
        current_settings["scheduler"]["main_day_monthly"] = day_main_monthly
    current_settings["scheduler"]["rot_enabled"] = False 
    if save_settings(current_settings): 
        logger.info(f"Schedule settings saved (Main En={is_enabled_main}, ROT cron effectively disabled).")
        return RedirectResponse(url=app.url_path_for('settings_page') + "?msg=schedule_settings_saved_apply_needed", status_code=303)
    else: 
        logger.error("Failed to save schedule settings.")
        return RedirectResponse(url=app.url_path_for('settings_page') + "?error=schedule_settings_save_failed", status_code=303)

@app.post("/apply-schedule", name="apply_schedule") 
async def apply_schedule(request: Request):
    logger.info("Apply schedule request. Executing scheduler_setup.py...")
    if not SCHEDULER_SETUP_SCRIPT_PATH.is_file(): logger.error(f"Scheduler script missing: {SCHEDULER_SETUP_SCRIPT_PATH}"); return RedirectResponse(url=app.url_path_for('settings_page') + "?error=scheduler_script_missing", status_code=303)
    try:
        py_exec = sys.executable ; cmd = [py_exec, str(SCHEDULER_SETUP_SCRIPT_PATH)]; logger.info(f"Executing: {' '.join(cmd)}")
        res = subprocess.run(cmd, capture_output=True, text=True, check=False, encoding='utf-8', cwd=SCRIPT_DIR)
        log_out = res.stdout.strip(); log_err = res.stderr.strip();
        if log_out: logger.info(f"Scheduler Setup Output:\n{log_out}")
        if log_err: logger.warning(f"Scheduler Setup Error Output:\n{log_err}")
        if res.returncode == 0: logger.info("scheduler_setup.py OK."); return RedirectResponse(url=app.url_path_for('settings_page') + "?msg=schedule_applied_successfully", status_code=303)
        else: logger.error(f"scheduler_setup.py failed: {res.returncode}."); err_part = log_err.splitlines()[-1] if log_err else "UnknownErr"; safe_err = re.sub(r'[^\w\s-]', '', err_part)[:50]; return RedirectResponse(url=app.url_path_for('settings_page') + f"?error=schedule_apply_failed_{safe_err}", status_code=303)
    except FileNotFoundError: logger.error(f"Failed scheduler_setup.py: Python '{py_exec}' or script not found."); return RedirectResponse(url=app.url_path_for('settings_page') + f"?error=schedule_apply_file_not_found", status_code=303)
    except Exception as e: logger.error(f"Failed scheduler_setup.py: {e}", exc_info=True); return RedirectResponse(url=app.url_path_for('settings_page') + f"?error=schedule_apply_exception_{type(e).__name__}", status_code=303)

@app.post("/run-cleanup", name="run_cleanup_task") 
async def run_cleanup_task(request: Request):
    logger.info("Manual cleanup task triggered.")
    try: cleanup_old_filtered_results(); logger.info("Manual cleanup task finished."); return RedirectResponse(url=app.url_path_for('settings_page') + "?msg=cleanup_finished_successfully", status_code=303)
    except Exception as e: logger.error(f"Manual cleanup error: {e}", exc_info=True); return RedirectResponse(url=app.url_path_for('settings_page') + f"?error=cleanup_failed_{type(e).__name__}", status_code=303)

@app.get("/view-logs", name="view_logs_page", response_class=HTMLResponse)
async def view_logs_page(request: Request, site_log: Optional[str] = None, log_type: str = "regular"):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine error.") # Added status_code

    log_lines_to_show = 200  # Number of lines to show from the end of the log
    
    # Controller log path
    controller_log_filename = "site_controller.log"
    # MODIFIED: Use new base log directory
    controller_log_path = LOGS_BASE_DIR / controller_log_filename

    def get_last_log_lines(log_path_func: Path, num_lines_func: int) -> Optional[str]:
        if not log_path_func.is_file():
            logger.warning(f"Log file not found: {log_path_func}")
            return f"[Log file not found: {log_path_func.name}]"
        try:
            # This method of reading last N lines is okay for moderately sized files.
            # For very large files, more optimized methods exist, but this is generally fine.
            with open(log_path_func, 'rb') as f_log: # Read as bytes
                f_log.seek(0, os.SEEK_END)
                end_byte_log = f_log.tell()
                lines_to_go_log = num_lines_func + 1 # Read one extra to handle potential partial last line
                block_size_log = 1024
                bytes_read_log = 0
                blocks_log = []
                while lines_to_go_log > 0 and bytes_read_log < end_byte_log:
                    read_size_log = min(block_size_log, end_byte_log - bytes_read_log)
                    f_log.seek(-(bytes_read_log + read_size_log), os.SEEK_END)
                    block_data_log = f_log.read(read_size_log)
                    blocks_log.append(block_data_log)
                    bytes_read_log += read_size_log
                    # More robust line counting for different OS line endings
                    lines_to_go_log -= block_data_log.count(b'\n')
                all_bytes_log = b''.join(reversed(blocks_log))
                all_text_log = all_bytes_log.decode('utf-8', errors='replace')
                lines_log = all_text_log.splitlines()
                # Get the actual last N lines, ensuring not to go out of bounds
                last_lines_log = lines_log[-num_lines_func:]
                return "\n".join(last_lines_log)
        except Exception as e_log_read:
            logger.error(f"Error reading log {log_path_func}: {e_log_read}", exc_info=True)
            return f"[Error reading log '{log_path_func.name}': {e_log_read}]"

    controller_log_content_val = get_last_log_lines(controller_log_path, log_lines_to_show)

    # MODIFIED: Paths for regular and ROT logs
    regular_scraper_logs_dir = LOGS_BASE_DIR / "regular_scraper"
    rot_worker_logs_base_dir = LOGS_BASE_DIR / "rot_worker" # This dir contains site_key subdirs

    regular_site_log_files: List[str] = []
    if regular_scraper_logs_dir.is_dir():
        regular_site_log_files = sorted([
            p.name for p in regular_scraper_logs_dir.glob("scrape_*.log") if p.is_file()
        ])
    else:
        logger.warning(f"Regular scraper log directory not found: {regular_scraper_logs_dir}")

    rot_site_log_files: List[str] = []
    if rot_worker_logs_base_dir.is_dir():
        # ROT logs are now nested under site_key, so we list filenames differently
        # For the dropdown, we might want to present "SITE_KEY/filename.log" or just filename if unique
        # For simplicity here, we'll list all log files within all site_key subdirs.
        # The template might need adjustment if you want to group by site_key in dropdown.
        # Or, better, list site_keys, then let user pick a site, then list logs for that site.
        # For now, flat list of log names (might have duplicates if filename pattern is same across sites):
        # This is a placeholder - ideally, you'd list site_keys first for ROT.
        # For a simpler dropdown for now, we'll collect all log names.
        temp_rot_logs = []
        for site_key_dir in rot_worker_logs_base_dir.iterdir():
            if site_key_dir.is_dir():
                for log_file in site_key_dir.glob("worker_*.log"):
                    # Prepend site_key to filename for uniqueness in dropdown if needed
                    temp_rot_logs.append(f"{site_key_dir.name}/{log_file.name}")
        rot_site_log_files = sorted(temp_rot_logs)

    else:
        logger.warning(f"ROT worker log base directory not found: {rot_worker_logs_base_dir}")

    selected_site_log_filename_val: Optional[str] = None
    selected_site_log_content_val: Optional[str] = None
    
    # Determine the correct base directory for the selected log type
    current_log_dir_for_selected: Optional[Path] = None
    actual_log_file_name_to_read: Optional[str] = None

    if site_log: # site_log might now be "SITE_KEY/worker_....log" for ROT
        if log_type == "regular" and regular_scraper_logs_dir.is_dir() and (regular_scraper_logs_dir / site_log).is_file():
            current_log_dir_for_selected = regular_scraper_logs_dir
            actual_log_file_name_to_read = site_log
        elif log_type == "rot" and rot_worker_logs_base_dir.is_dir():
            # site_log for ROT is expected to be "SITE_KEY/filename.log"
            if '/' in site_log:
                # path_parts = site_log.split('/', 1) # site_key, actual_filename
                # potential_path = rot_worker_logs_base_dir / path_parts[0] / path_parts[1]
                potential_path = rot_worker_logs_base_dir / site_log # site_log is already SITE_KEY/filename
                if potential_path.is_file():
                    # current_log_dir_for_selected will be the specific site_key dir
                    current_log_dir_for_selected = potential_path.parent
                    actual_log_file_name_to_read = potential_path.name
            else: # Legacy: if site_log is just filename, try finding it (less robust)
                for sk_dir in rot_worker_logs_base_dir.iterdir():
                    if sk_dir.is_dir() and (sk_dir / site_log).is_file():
                        current_log_dir_for_selected = sk_dir
                        actual_log_file_name_to_read = site_log
                        break
        
        if current_log_dir_for_selected and actual_log_file_name_to_read:
            selected_site_log_filename_val = site_log # Keep the full "SITE_KEY/filename" for display consistency
            selected_site_log_content_val = get_last_log_lines(current_log_dir_for_selected / actual_log_file_name_to_read, log_lines_to_show)
        elif site_log: # site_log was provided but not found
             logger.warning(f"Requested site_log '{site_log}' of type '{log_type}' not found in expected location.")
             selected_site_log_content_val = f"[Log file '{site_log}' (Type: {log_type}) not found]"


    # Assuming the template `view_logs.html` is expecting `current_log_type_filter`
    # The original code was truncated, I'm adding it back.
    return templates.TemplateResponse(
        "view_logs.html",
        {
            "request": request,
            "controller_log_filename": controller_log_filename,
            "controller_log_content": controller_log_content_val,
            "regular_site_log_files": regular_site_log_files, # List of "scrape_SITE.log"
            "rot_site_log_files": rot_site_log_files,         # List of "SITE_KEY/worker_RUNID_DATE.log"
            "selected_site_log_filename": selected_site_log_filename_val, # Could be "scrape_SITE.log" or "SITE_KEY/worker..."
            "selected_site_log_content": selected_site_log_content_val,
            "current_log_type_filter": log_type # This was the truncated part
        }
    )

@app.get("/download-log/{log_type}/{path_param_site_key_or_filename:path}", name="download_log_typed")
async def download_log_typed(log_type: str, path_param_site_key_or_filename: str):
    # path_param_site_key_or_filename will be just 'filename' for controller and regular,
    # but 'SITE_KEY/filename' for ROT logs.
    
    logger.info(f"Log download request: Type='{log_type}', PathParam='{path_param_site_key_or_filename}'")

    log_path_to_serve: Optional[Path] = None
    actual_filename_for_download: str = path_param_site_key_or_filename # Default

    if log_type == "controller" and path_param_site_key_or_filename == "site_controller.log":
        log_path_to_serve = LOGS_BASE_DIR / "site_controller.log"
        actual_filename_for_download = "site_controller.log"
    elif log_type == "regular" and path_param_site_key_or_filename.startswith("scrape_") and path_param_site_key_or_filename.endswith(".log"):
        log_path_to_serve = LOGS_BASE_DIR / "regular_scraper" / path_param_site_key_or_filename
        actual_filename_for_download = path_param_site_key_or_filename
    elif log_type == "rot":
        # path_param_site_key_or_filename is expected to be "SITE_KEY/worker_RUNID_DATE.log"
        # The Path object will handle joining these correctly.
        # For example, if path_param is "CPPP/worker_abc.log", then
        # LOGS_BASE_DIR / "rot_worker" / "CPPP/worker_abc.log" becomes
        # /path/to/LOGS/rot_worker/CPPP/worker_abc.log
        log_path_to_serve = LOGS_BASE_DIR / "rot_worker" / path_param_site_key_or_filename
        actual_filename_for_download = Path(path_param_site_key_or_filename).name # Just the "worker_RUNID_DATE.log" part
    else:
        logger.warning(f"Invalid log type or filename pattern for download: Type='{log_type}', PathParam='{path_param_site_key_or_filename}'")
        raise HTTPException(status_code=400, detail="Invalid log type or filename for download.")

    if log_path_to_serve is None: # Should be caught by the else above, but as a safeguard
        logger.error(f"Log path could not be determined for download: Type='{log_type}', PathParam='{path_param_site_key_or_filename}'")
        raise HTTPException(status_code=500, detail="Log path determination error.")

    resolved_log_path = log_path_to_serve.resolve()

    # Security Check: Ensure the resolved path is within the intended LOGS directory structure
    # This needs to be robust for different log types.
    allowed_base_for_log_type: Path
    if log_type == "controller":
        allowed_base_for_log_type = (LOGS_BASE_DIR).resolve()
    elif log_type == "regular":
        allowed_base_for_log_type = (LOGS_BASE_DIR / "regular_scraper").resolve()
    elif log_type == "rot":
        allowed_base_for_log_type = (LOGS_BASE_DIR / "rot_worker").resolve()
    else: # Should not be reached if initial checks are correct
        raise HTTPException(status_code=500, detail="Internal log security check error.")

    if not str(resolved_log_path).startswith(str(allowed_base_for_log_type)) or \
       allowed_base_for_log_type not in resolved_log_path.parents:
        logger.error(f"Path traversal attempt for log download: Type='{log_type}', PathParam='{path_param_site_key_or_filename}', Resolved='{resolved_log_path}', ExpectedBase='{allowed_base_for_log_type}'")
        raise HTTPException(status_code=403, detail="Forbidden: Access to this log path is not allowed.")

    if not resolved_log_path.is_file():
        logger.warning(f"Log file not found for download at: {resolved_log_path}")
        raise HTTPException(status_code=404, detail=f"Log file '{path_param_site_key_or_filename}' (Type: {log_type}) not found.")

    try:
        logger.info(f"Serving log download: {resolved_log_path} as {actual_filename_for_download}")
        return FileResponse(
            path=resolved_log_path,
            media_type='text/plain',
            filename=actual_filename_for_download # Serve with the simple filename
        )
    except Exception as e_serve_log:
        logger.error(f"Error serving log {resolved_log_path}: {e_serve_log}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error serving log file.")

@app.get("/download-rot-summary/{site_key}/{filename}", name="download_rot_summary_file")
async def download_rot_summary_file(site_key: str, filename: str):
    logger.info(f"Request to download ROT summary file: SiteKey='{site_key}', Filename='{filename}'")
    if not site_key or not filename:
        logger.warning("Download ROT summary: Missing site_key or filename.")
        raise HTTPException(status_code=400, detail="Site key and filename are required.")

    # REMOVE the local definition of _clean_path_component if it was here.
    # It should now be a global helper function.
    # def _clean_path_component(component: str) -> str: ... (This line should be deleted from inside this function)

    # Use the global helper function
    cleaned_site_key = _clean_path_component(site_key)
    cleaned_filename = _clean_path_component(filename)

    if not cleaned_site_key: # _clean_path_component returns "" for invalid/unsafe input
        logger.warning(f"Download ROT summary: Invalid site_key after cleaning. Original: '{site_key}'")
        raise HTTPException(status_code=400, detail="Invalid site key format.")
    if not cleaned_filename: # _clean_path_component returns "" for invalid/unsafe input
        logger.warning(f"Download ROT summary: Invalid filename after cleaning. Original: '{filename}'")
        raise HTTPException(status_code=400, detail="Invalid filename format.")

    # Uses the global constant ROT_DETAIL_HTMLS_DIR
    file_path = (ROT_DETAIL_HTMLS_DIR / cleaned_site_key / cleaned_filename).resolve()
    intended_base_for_site_files = (ROT_DETAIL_HTMLS_DIR / cleaned_site_key).resolve()

    # Path traversal check
    if not str(file_path).startswith(str(intended_base_for_site_files)) or \
       intended_base_for_site_files not in file_path.parents:
        logger.error(f"Download ROT summary: Path traversal attempt. "
                     f"Req Filename: '{filename}', Cleaned FN: '{cleaned_filename}', Site: '{cleaned_site_key}', "
                     f"Resolved Path: '{file_path}', Expected Base: '{intended_base_for_site_files}'")
        raise HTTPException(status_code=403, detail="Access forbidden.")

    if not file_path.is_file():
        logger.error(f"Download ROT summary: File not found at resolved path: {file_path}")
        raise HTTPException(status_code=404, detail=f"File '{cleaned_filename}' not found for site '{cleaned_site_key}'.")

    try:
        logger.info(f"Serving ROT summary file: {file_path}")
        media_type = "application/octet-stream" # Default
        file_extension = Path(cleaned_filename).suffix.lower()
        if file_extension == ".pdf":
            media_type = "application/pdf"
        elif file_extension in [".xls", ".xlsx"]:
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif file_extension == ".html" or file_extension == ".htm":
            media_type = "text/html"
        # Add more mime types if needed

        return FileResponse(path=file_path, filename=cleaned_filename, media_type=media_type)
    except Exception as e:
        logger.error(f"Error serving ROT summary file {file_path}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error serving file.")

@app.post("/run-scraper-now", name="run_scraper_now")
async def run_scraper_now(request: Request, scrape_type: str = Form("regular")):
    logger.info(f"Manual 'Run Scraper Now' ({scrape_type}) triggered via dashboard.")
    if scrape_type == "rot":
        logger.info("ROT scraping is manual via the dedicated 'Manual ROT Scrape' page. Ignoring general 'Fetch ROT'.")
        return RedirectResponse(url=app.url_path_for('homepage') + "?msg=manual_scrape_rot_is_manual_via_dedicated_page", status_code=303)

    scrape_type_to_run = "regular"
    message_key = "manual_scrape_regular_triggered"
    if scrape_type == "all":
        logger.info("'Fetch All' now only triggers REGULAR scrape. ROT is manual.")
        message_key = "manual_scrape_all_triggers_regular_only"

    # SITE_CONTROLLER_SCRIPT_PATH should be defined globally using PROJECT_ROOT
    if not SITE_CONTROLLER_SCRIPT_PATH.is_file():
        logger.error(f"Site controller script not found: {SITE_CONTROLLER_SCRIPT_PATH}")
        return RedirectResponse(url=app.url_path_for('homepage') + "?error=controller_script_missing", status_code=303)
    try:
        python_executable_val = sys.executable # Correctly gets the venv python if dashboard runs in venv
        command_list = [python_executable_val, str(SITE_CONTROLLER_SCRIPT_PATH), "--type", scrape_type_to_run]
        logger.info(f"Executing command in background: {' '.join(command_list)}")
        
        # MODIFIED: Use PROJECT_ROOT for cwd
        process_bg = subprocess.Popen(command_list, cwd=PROJECT_ROOT,
                                      stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                      text=True, encoding='utf-8')
        
        logger.info(f"Launched site_controller.py (Type: {scrape_type_to_run}) BG (PID: {process_bg.pid}). Check logs.")
        return RedirectResponse(url=app.url_path_for('homepage') + f"?msg={message_key}", status_code=303)
    except FileNotFoundError:
        # This specific exception might not be hit if python_executable_val is from sys.executable
        # and SITE_CONTROLLER_SCRIPT_PATH check is done above. But good to keep.
        logger.error(f"Failed execute site_controller.py: Python '{python_executable_val}' or script missing.")
        return RedirectResponse(url=app.url_path_for('homepage') + f"?error=manual_scrape_file_not_found_{scrape_type_to_run}", status_code=303)
    except Exception as e_launch:
        logger.error(f"Failed launch site_controller.py (Type: {scrape_type_to_run}): {e_launch}", exc_info=True)
        return RedirectResponse(url=app.url_path_for('homepage') + f"?error=manual_scrape_launch_failed_{scrape_type_to_run}_{type(e_launch).__name__}", status_code=303)
# === AI Analysis Backend Endpoints ===

@app.post("/rot-prepare-analysis-data", name="rot_prepare_analysis_data")
async def rot_prepare_analysis_data_route():
    logger.info("AI Data Preparation: Received request to collate ROT data.")
    if not filter_engine_available:
        logger.error("AI Data Preparation: Filter engine components are not available.")
        raise HTTPException(status_code=500, detail="Core data processing components (filter_engine) are missing.")
    try:
        merged_count, latest_final_rot_file_path = await _globally_merge_rot_site_files() # This now uses new paths internally

        if latest_final_rot_file_path is None or not latest_final_rot_file_path.is_file():
            message_detail = "No site-specific ROT data found to merge globally."
            if merged_count == 0 and latest_final_rot_file_path is None:
                 logger.warning(f"AI Data Preparation: {message_detail}")
            else:
                 message_detail = "Global ROT data merge failed or produced no output file."
                 logger.error(f"AI Data Preparation: {message_detail}")
            # MODIFIED: Update error message to reflect new input directory for _globally_merge_rot_site_files
            raise HTTPException(status_code=404, detail=f"{message_detail} Ensure 'Merged_ROT_SITEKEY_DATE.txt' files exist in '{ROT_MERGED_SITE_SPECIFIC_DIR}'.")

        logger.info(f"AI Data Preparation: Using globally merged source file: {latest_final_rot_file_path.name} from {latest_final_rot_file_path.parent}")
        tagged_blocks = parse_tender_blocks_from_tagged_file(latest_final_rot_file_path)
        collated_rot_tenders: List[Dict[str, Any]] = []
        if tagged_blocks:
            for block_text in tagged_blocks:
                try:
                    tender_info = extract_tender_info_from_tagged_block(block_text)
                    if tender_info.get("data_type") == "rot": # Ensure only ROT data is collated
                        collated_rot_tenders.append(tender_info)
                except Exception as e_parse_block:
                    logger.error(f"AI Data Preparation: Error processing a tender block: {e_parse_block}", exc_info=False)
        
        # AI_ANALYSIS_DATA_FILE should now be using ROT_AI_ANALYSIS_DIR from the top config
        AI_ANALYSIS_DATA_FILE.parent.mkdir(parents=True, exist_ok=True) # Ensure parent dir exists
        with open(AI_ANALYSIS_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(collated_rot_tenders, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"AI Data Preparation: Collated {len(collated_rot_tenders)} ROT items from '{latest_final_rot_file_path.name}' into {AI_ANALYSIS_DATA_FILE}")
        
        # Make file_path relative to PROJECT_ROOT for the response
        # SCRIPT_DIR should have been replaced by PROJECT_ROOT in the global constants section
        relative_ai_file_path = str(AI_ANALYSIS_DATA_FILE.relative_to(PROJECT_ROOT)) # MODIFIED

        return JSONResponse(
            content={
                "message": f"ROT data collated from '{latest_final_rot_file_path.name}'. {len(collated_rot_tenders)} items processed.",
                "file_path": relative_ai_file_path # MODIFIED
            }, status_code=status.HTTP_200_OK )
    except HTTPException as http_exc:
        # Log the specific HTTP exception details if needed, then re-raise
        logger.error(f"AI Data Preparation HTTP Error from within route: {http_exc.status_code} - {http_exc.detail}")
        raise http_exc
    except Exception as e:
        logger.error(f"AI Data Preparation Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to prepare ROT data: {type(e).__name__}")

# AI_ANALYSIS_DATA_FILE should be defined globally as:
# AI_ANALYSIS_DATA_FILE = ROT_AI_ANALYSIS_DIR / "ai_analysis_data.json"
# And ROT_AI_ANALYSIS_DIR should be:
# ROT_AI_ANALYSIS_DIR = SITE_DATA_ROOT / "ROT" / "AI_Analysis"
# (These definitions were part of the Step 3 CONFIGURATION block update)

@app.get("/get-collated-rot-data-json", name="get_collated_rot_data_json")
async def get_collated_rot_data_json_route():
    logger.info("AI Data Request: Serving collated ROT data.")
    # This function relies on AI_ANALYSIS_DATA_FILE being correctly defined globally
    # with the new path structure.
    if not AI_ANALYSIS_DATA_FILE.is_file():
        logger.warning(f"AI Data Request: Collated file not found: {AI_ANALYSIS_DATA_FILE}")
        # The error message to the user is fine as is.
        raise HTTPException(status_code=404, detail="Collated ROT data not prepared. Run 'Prepare Data' first.")
    try:
        with open(AI_ANALYSIS_DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return JSONResponse(content=data)
    except Exception as e:
        logger.error(f"AI Data Request: Error serving collated data from {AI_ANALYSIS_DATA_FILE}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to serve collated ROT data.")

class AIPromptRequest(BaseModel): # This model is fine, no path changes needed.
    model: Optional[str] = None
    prompt: str

@app.post("/proxy-ai-chat", name="proxy_ai_chat")
async def proxy_ai_chat_route(ai_request: AIPromptRequest):
    logger.info(f"AI Proxy: Chat request. Model: {ai_request.model or DEFAULT_OLLAMA_MODEL_ID_CONFIG}")
    target_model = ai_request.model if ai_request.model else DEFAULT_OLLAMA_MODEL_ID_CONFIG
    openwebui_chat_endpoint = f"{OPENWEBUI_API_BASE_URL_CONFIG.rstrip('/')}/api/chat/completions" 
    payload = {"model": target_model, "messages": [{"role": "user", "content": ai_request.prompt}], "stream": False}
    headers = {'Content-Type': 'application/json'}
    if OPENWEBUI_API_KEY_CONFIG: 
        headers['Authorization'] = f'Bearer {OPENWEBUI_API_KEY_CONFIG}'
    logger.debug(f"AI Proxy: Sending to OpenWebUI. Endpoint: {openwebui_chat_endpoint}, Model: {target_model}")
    try:
        async with httpx.AsyncClient(timeout=180.0) as client:
            response = await client.post(openwebui_chat_endpoint, json=payload, headers=headers)
            response.raise_for_status() 
            ai_response_data = response.json()
            logger.info(f"AI Proxy: Received response from OpenWebUI (Status: {response.status_code}).")
            return JSONResponse(content=ai_response_data)
    except httpx.HTTPStatusError as exc:
        error_content_for_logging = "Raw text or unknown structure from AI service error."
        error_detail_for_fastapi = "AI service returned an error."
        try:
            error_content_json = exc.response.json()
            if isinstance(error_content_json, dict) and "error" in error_content_json:
                if isinstance(error_content_json["error"], dict) and "message" in error_content_json["error"]:
                    error_content_for_logging = error_content_json["error"]["message"]
                    error_detail_for_fastapi = error_content_json["error"]["message"]
                elif isinstance(error_content_json["error"], str):
                    error_content_for_logging = error_content_json["error"]
                    error_detail_for_fastapi = error_content_json["error"]
                else:
                    error_content_for_logging = str(error_content_json) 
                    error_detail_for_fastapi = str(error_content_json)
            elif isinstance(error_content_json, dict) and "detail" in error_content_json: 
                error_content_for_logging = error_content_json["detail"]
                error_detail_for_fastapi = error_content_json["detail"]
            else: 
                error_content_for_logging = str(error_content_json)
                error_detail_for_fastapi = "Error detail not in expected format from AI service."
        except json.JSONDecodeError: 
            error_content_for_logging = exc.response.text
            error_detail_for_fastapi = exc.response.text[:200] 
        logger.error(f"AI Proxy: HTTP Error from OpenWebUI: {exc.response.status_code} - Detail: {error_content_for_logging}")
        raise HTTPException(status_code=exc.response.status_code, detail=str(error_detail_for_fastapi))
    except httpx.RequestError as exc:
        logger.error(f"AI Proxy: Request Error to OpenWebUI: {exc}")
        raise HTTPException(status_code=503, detail=f"AI service connection error: {type(exc).__name__}")
    except Exception as e:
        logger.error(f"AI Proxy: Unexpected error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"AI proxy unexpected error: {type(e).__name__}")

# === Endpoints for scraper2.py CAPTCHA Interaction ===
@app.get("/solve-captcha/{run_id}", name="solve_captcha_page", response_class=HTMLResponse)
async def solve_captcha_page_route(request: Request, run_id: str):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine not configured.")
    if not re.match(r"^[a-zA-Z0-9_.-]+$", run_id) or ".." in run_id:
        raise HTTPException(status_code=400, detail="Invalid run_id format.")
    return templates.TemplateResponse("solve_captcha.html", {"request": request, "run_id": run_id})

@app.get("/get-captcha-status/{run_id}", name="get_captcha_status")
async def get_captcha_status_route(run_id: str):
    if not re.match(r"^[a-zA-Z0-9_.-]+$", run_id) or ".." in run_id:
        logger.warning(f"Dashboard: Invalid run_id format in get_captcha_status: {run_id}")
        raise HTTPException(status_code=400, detail="Invalid run_id format.")

    safe_run_id_name = Path(run_id).name
    # Use new constant for worker communication directory
    run_dir = TEMP_WORKER_RUNS_DIR / safe_run_id_name # MODIFIED

    status_file = run_dir / "status.txt"
    captcha_b64_file = run_dir / "captcha.b64"

    response_data = {
        "run_id": run_id,
        "status": "pending_worker_init",
        "message": "Waiting for worker to initialize and generate CAPTCHA...",
        "image_data": None
    }

    if not run_dir.is_dir():
        response_data["status"] = "error_run_dir_not_found"
        response_data["message"] = "Error: Worker run directory not found."
        logger.warning(f"Dashboard: Run directory '{run_dir}' not found for run_id '{run_id}'.")
        return JSONResponse(content=response_data, status_code=status.HTTP_404_NOT_FOUND)

    if status_file.is_file():
        try:
            current_status_msg = status_file.read_text(encoding='utf-8').strip()
            response_data["message"] = f"Worker status: {current_status_msg}"

            if current_status_msg == "WAITING_CAPTCHA":
                response_data["status"] = "ready_for_captcha"
                response_data["message"] = "CAPTCHA image ready. Please view and solve."
                if captcha_b64_file.is_file():
                    try:
                        b64_data = captcha_b64_file.read_text(encoding='utf-8').strip()
                        if b64_data:
                            response_data["image_data"] = f"data:image/png;base64,{b64_data}"
                        else:
                            # Worker might still be writing, or file is genuinely empty.
                            response_data["status"] = "processing_captcha_file";
                            response_data["message"] = "Worker is generating CAPTCHA image data (file found but empty)..."
                            logger.warning(f"Dashboard: captcha.b64 for run_id '{run_id}' is empty.")
                    except Exception as e_read_b64:
                        response_data["status"] = "error_reading_captcha_file";
                        response_data["message"] = "Error: Could not read CAPTCHA image data."
                        logger.error(f"Dashboard: Error reading captcha.b64 for '{run_id}': {e_read_b64}")
                else:
                    response_data["status"] = "processing_captcha_file";
                    response_data["message"] = "Worker is preparing CAPTCHA image file (file not yet found)..."
            # Refined status checking for clarity and to catch more specific worker error states
            elif current_status_msg.startswith("ERROR_") or current_status_msg == "TIMEOUT_CAPTCHA_INPUT":
                response_data["status"] = "worker_error"; # Generic error for UI
                # Message already contains specific error from status.txt
            elif current_status_msg.startswith("FINISHED_"):
                response_data["status"] = "worker_finished"; # Generic finished for UI
                # Message already contains specific finish status
            elif current_status_msg in ["INITIATED_BY_DASHBOARD", "WORKER_STARTED", "FETCHING_CAPTCHA",
                                        "PROCESSING_WITH_CAPTCHA", "SCRAPING_RESULTS"] \
                 or current_status_msg.startswith("PROCESSING_LIST_PAGE_"):
                 response_data["status"] = "worker_processing";
            # No 'else' needed here, if status is unrecognized, it keeps default "pending_worker_init"
            # or the message just reflects the unknown status.
        except Exception as e_read_status:
            response_data["status"] = "error_reading_status_file";
            response_data["message"] = "Error: Could not read worker status file."
            logger.error(f"Dashboard: Error reading status.txt for '{run_id}': {e_read_status}")
    else:
        response_data["message"] = "Worker process is initializing (status file not yet created)..."
        # Keeps "pending_worker_init" status

    return JSONResponse(content=response_data)

@app.post("/submit-captcha-answer/{run_id}", name="submit_captcha_answer")
async def submit_captcha_answer_route(run_id: str, captcha_text: str = Form(...)):
    if not re.match(r"^[a-zA-Z0-9_.-]+$", run_id) or ".." in run_id:
        logger.warning(f"Dashboard: Invalid run_id format in submit_captcha_answer: {run_id}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid run_id format.")

    logger.info(f"Dashboard: Received CAPTCHA answer for run_id '{run_id}': '{captcha_text}'")

    safe_run_id_name = Path(run_id).name
    # Use new constant for worker communication directory
    run_dir = TEMP_WORKER_RUNS_DIR / safe_run_id_name # MODIFIED
    answer_file = run_dir / "answer.txt"
    status_file = run_dir / "status.txt"

    if not run_dir.is_dir():
        logger.error(f"Dashboard: Run directory '{run_dir}' not found for run_id '{run_id}'.")
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"detail": f"Worker run '{run_id}' not found."}
        )

    try:
        if status_file.is_file():
            current_worker_status = status_file.read_text(encoding='utf-8').strip().upper()
            # Check for specific terminal prefixes for more robustness
            terminal_prefixes = ["FINISHED_", "ERROR_", "TIMEOUT_"]
            if any(current_worker_status.startswith(prefix) for prefix in terminal_prefixes):
                logger.warning(f"Dashboard: Attempt to submit CAPTCHA for terminated run '{run_id}'. Status: '{current_worker_status}'")
                return JSONResponse(
                    status_code=status.HTTP_409_CONFLICT,
                    content={"detail": f"Worker run '{run_id}' is already in a terminal state ({current_worker_status})."}
                )

        answer_file.write_text(captcha_text, encoding='utf-8')
        # write_dashboard_status is a helper defined in dashboard.py
        write_dashboard_status(run_dir, "CAPTCHA_ANSWER_SUBMITTED_BY_DASHBOARD")
        logger.info(f"Dashboard: Wrote answer to '{answer_file}' for run_id '{run_id}'.")

        return JSONResponse(
            content={"message": f"CAPTCHA answer for run_id '{run_id}' submitted.", "run_id": run_id},
            status_code=status.HTTP_200_OK
        )
    except Exception as e:
        logger.error(f"Dashboard: Error writing answer file for run_id '{run_id}': {e}", exc_info=True)
        # Attempt to update status to reflect this dashboard-side error if possible
        if run_dir.is_dir(): # Check again in case it was deleted or inaccessible before
            write_dashboard_status(run_dir, f"ERROR_DASHBOARD_SUBMITTING_ANSWER_{type(e).__name__}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"detail": "Failed to submit CAPTCHA answer to worker."}
        )

# In dashboard.py

@app.post("/rot/initiate-rot-worker/{site_key}", name="initiate_rot_worker")
async def initiate_rot_worker_endpoint(request: Request, site_key: str, tender_status_value: str = Form(...)):
    logger.info(f"Dashboard: Initiating ROT Worker for Site='{site_key}', Status='{tender_status_value}'")
    if not all([site_key, tender_status_value]):
        logger.error("Dashboard: Missing site_key or tender_status_value for worker initiation.")
        raise HTTPException(status_code=400, detail="Site key and tender status value are required.")

    worker_script_name = "headless_rot_worker.py"
    worker_script_path = PROJECT_ROOT / worker_script_name

    if not worker_script_path.is_file():
        logger.error(f"Dashboard: ROT Worker script '{worker_script_name}' not found at {worker_script_path}.")
        raise HTTPException(status_code=500, detail=f"ROT worker script '{worker_script_name}' component is missing on the server.")

    # --- Create run_id and run_dir FIRST ---
    # This is so we can write error status to it if subsequent steps fail
    cleaned_site_key_for_id = re.sub(r'[^\w\-]+', '_', site_key)
    run_id = f"rot_worker_{cleaned_site_key_for_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    run_dir = TEMP_WORKER_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True) # Create dir before trying to write status

    try:
        python_exec = sys.executable

        # --- Get site-specific URLs ---
        rot_site_urls = get_rot_site_config(site_key) # Must be defined and working
        if not rot_site_urls or not rot_site_urls.get("main_search_url") or not rot_site_urls.get("base_url"):
            err_msg_no_url = f"Configuration error: Could not get URLs for ROT site '{site_key}'."
            logger.error(f"Dashboard: {err_msg_no_url}")
            write_dashboard_status(run_dir, f"ERROR_CONFIG_NO_URLS_FOR_SITE_{site_key}") # Write error to run_dir
            raise HTTPException(status_code=500, detail=err_msg_no_url)

        # --- Prepare worker settings JSON ---
        current_settings = load_settings() # Must be defined and working
        gs_global_settings = current_settings.get("global_scraper_settings", {})
        gs_rot_settings = gs_global_settings.get("rot_scrape_limits", {})
        gs_rot_concurrency = gs_global_settings.get("rot_concurrency", {})
        gs_rot_timeouts = gs_global_settings.get("rot_timeouts", {})

        # Fallback to DEFAULT_SETTINGS (defined globally in dashboard.py)
        default_rot_limits = DEFAULT_SETTINGS["global_scraper_settings"]["rot_scrape_limits"]
        default_rot_concurrency = DEFAULT_SETTINGS["global_scraper_settings"]["rot_concurrency"]
        default_rot_timeouts = DEFAULT_SETTINGS["global_scraper_settings"]["rot_timeouts"]

        worker_run_params = {
            "max_list_pages": gs_rot_settings.get("max_list_pages", default_rot_limits["max_list_pages"]),
            "detail_concurrency": gs_rot_concurrency.get("detail_processing", default_rot_concurrency["detail_processing"]),
            "pagination_wait": gs_rot_timeouts.get("pagination_load_wait", default_rot_timeouts["pagination_load_wait"]),
            "page_load_timeout": gs_rot_timeouts.get("page_load", default_rot_timeouts["page_load"]),
            "detail_page_timeout": gs_rot_timeouts.get("detail_page", default_rot_timeouts["detail_page"]),
            "popup_page_timeout": gs_rot_timeouts.get("detail_page", default_rot_timeouts["detail_page"]), # Using detail_page as fallback
            "element_timeout": gs_rot_timeouts.get("element_wait", default_rot_timeouts["element_wait"]),
            "post_submit_timeout": gs_rot_timeouts.get("post_submit", default_rot_timeouts["post_submit"])
        }
        worker_settings_json_str = json.dumps(worker_run_params)
        # --- End Prepare worker settings ---

        # Initialize status file (now that we have run_id and run_dir)
        write_dashboard_status(run_dir, "INITIATED_BY_DASHBOARD") # Initial status
        logger.info(f"Dashboard: Created run directory '{run_dir}' and status file for Run ID: {run_id}")

        command = [
            python_exec, str(worker_script_path),
            "--site_key", site_key,
            "--run_id", run_id,
            "--site_url", rot_site_urls["main_search_url"],   # CORRECTED
            "--base_url", rot_site_urls["base_url"],         # CORRECTED
            "--tender_status", tender_status_value,
            "--settings_json", worker_settings_json_str    # CORRECTED
        ]

        logger.info(f"Dashboard: Launching ROT Worker command: {' '.join(command)}")
        process = subprocess.Popen(command, cwd=PROJECT_ROOT,
                                   stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                   text=True, encoding='utf-8')
        logger.info(f"Dashboard: Launched background ROT Worker for {site_key} (PID: {process.pid}). Run ID: {run_id}")

        response_content = {
            "message": f"ROT Worker ({worker_script_name}) initiated for site '{site_key}'. Polling for CAPTCHA image will begin.",
            "run_id": run_id,
            "site_key": site_key
        }
        return JSONResponse(status_code=status.HTTP_202_ACCEPTED, content=response_content)
    except Exception as e:
        # This generic except block catches errors after run_dir is created
        # (e.g., during Popen, or if get_rot_site_config/load_settings raised something unexpected)
        error_type_name = type(e).__name__
        logger.error(f"Dashboard: Failed to launch/prepare worker '{worker_script_name}' for site '{site_key}'. Error: {error_type_name} - {e}", exc_info=True)
        write_dashboard_status(run_dir, f"ERROR_LAUNCHING_WORKER_{error_type_name}")
        raise HTTPException(status_code=500, detail=f"Failed to start ROT worker for {site_key}: {error_type_name}")

@app.get("/rot-summary-detail/{site_key}/{filename}", name="view_rot_summary_detail", response_class=HTMLResponse)
async def view_rot_summary_detail_route(request: Request, site_key: str, filename: str):
    if not templates:
        raise HTTPException(status_code=503, detail="Template engine error: Templates not initialized.")

    logger.info(f"Request for parsed ROT summary: SiteKey='{site_key}', Filename='{filename}'")
    cleaned_site_key = _clean_path_component(site_key) # For path construction
    cleaned_filename = _clean_path_component(filename)

    if not cleaned_site_key: raise HTTPException(status_code=400, detail="Invalid site key format provided.")
    if not cleaned_filename: raise HTTPException(status_code=400, detail="Invalid filename format provided.")

    html_file_path = ROT_DETAIL_HTMLS_DIR / cleaned_site_key / cleaned_filename

    if not html_file_path.is_file():
        logger.warning(f"Requested ROT summary HTML not found at: {html_file_path}")
        return templates.TemplateResponse("error.html", {
            "request": request,
            "error": f"The ROT summary file '{cleaned_filename}' for site '{cleaned_site_key}' was not found."
        }, status_code=404)

    try:
        # Pass the original site_key (with spaces if any) for potential use in title setting
        parsed_content = parse_rot_summary_html(html_file_path, site_key_original=site_key)
    except Exception as e_parse:
        logger.error(f"Error during parse_rot_summary_html for {html_file_path}: {e_parse}", exc_info=True)
        return templates.TemplateResponse("error.html", {
            "request": request,
            "error": f"Error parsing '{cleaned_filename}'. Check server logs."
        }, status_code=500)

    tender_id_from_filename = "UnknownTenderID"
    fn_match = re.match(r"ROT_.*?_(.+?)_\d{8}_\d{6}_StageSummary\.html", cleaned_filename)
    if fn_match: tender_id_from_filename = fn_match.group(1)
    
    # tender_list_info is now less critical as key_details are not separately displayed as primary
    tender_list_info = {
        "rot_tender_id": tender_id_from_filename, # Primarily use ID from filename
        "rot_title_ref": parsed_content.get("page_title", "Summary Details")
    }

    logger.info(f"Rendering parsed ROT summary for: {html_file_path}")
    return templates.TemplateResponse(
        "rot_summary_detail.html",
        {
            "request": request,
            "site_key": site_key, # Pass original site key for display
            "filename": cleaned_filename,
            "parsed_data": parsed_content,
            "tender_list_info": tender_list_info,
        }
    )

@app.post("/trigger-rot-consolidation-from-worker", name="trigger_rot_consolidation_from_worker")
async def trigger_rot_consolidation_from_worker_route():
    logger.info("Dashboard: Received request from a worker to trigger ROT data consolidation.")
    try:
        # _globally_merge_rot_site_files is already defined and uses new ROT paths
        merged_count, merged_file_path = await _globally_merge_rot_site_files(
            input_site_specific_merged_dir=ROT_MERGED_SITE_SPECIFIC_DIR,
            output_final_global_dir=ROT_FINAL_GLOBAL_MERGED_DIR,
            data_type="rot"
        )
        if merged_file_path and merged_file_path.is_file():
            msg = f"ROT data consolidation triggered by worker completed. {merged_count} blocks merged to {merged_file_path.name}."
            logger.info(msg)
            return {"message": msg, "file": str(merged_file_path)}
        elif merged_count == 0 and merged_file_path is None:
             msg = "ROT data consolidation triggered by worker found no new site-specific files to merge."
             logger.info(msg)
             return {"message": msg, "file": None}
        else:
            msg = "ROT data consolidation triggered by worker ran but did not produce an output file or reported an issue."
            logger.warning(msg)
            raise HTTPException(status_code=500, detail=msg)
    except Exception as e:
        logger.error(f"Error during worker-triggered ROT consolidation: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error during ROT consolidation: {str(e)}")

# --- END OF dashboard.py ---
